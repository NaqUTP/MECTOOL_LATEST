# streamlit_app.py
# MEC TOOL – PETRONAS themed UI, professional & minimalistic
# Author: Ahmad Naquib Syahmee Masror
# Date: 2026-03-09

import io
import re
import json
import base64
import warnings
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import pandas as pd
from PIL import Image
import streamlit as st

warnings.filterwarnings("ignore", category=UserWarning)

try:
    from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
except Exception:
    st.error("Missing streamlit-aggrid. Install: pip install streamlit-aggrid")
    st.stop()
try:
    import plotly.express as px
    import plotly.graph_objects as go
except Exception:
    st.error("Missing plotly. Install: pip install plotly")
    st.stop()

# ─────────────────────────────────────────────────
# Page config – must be FIRST Streamlit command
_logo_path = Path(__file__).parent / "mec-tool-logo.png"
_page_icon = Image.open(_logo_path) if _logo_path.exists() else "⛽"
st.set_page_config(
    page_title="MEC TOOL | PETRONAS",
    page_icon=_page_icon,
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────────
# PETRONAS Design System
PETRONAS = {
    "teal":       "#00A19C",   # Primary
    "teal_dark":  "#007A76",   # Hover / Dark
    "teal_deep":  "#004D4B",   # Deep bg accent
    "teal_light": "#E6F7F7",   # Tinted bg
    "gold":       "#C8A951",   # Accent / highlight
    "navy":       "#0A1628",   # Dark bg
    "navy_mid":   "#0F2040",   # Cards dark
    "slate":      "#1E3A5F",   # Sidebar
    "white":      "#FFFFFF",
    "off_white":  "#F4F7F9",
    "grey_100":   "#EDF1F5",
    "grey_300":   "#C2CDD8",
    "grey_500":   "#7A8EA3",
    "grey_700":   "#3D5068",
    "success":    "#22C55E",
    "warning":    "#F59E0B",
    "danger":     "#EF4444",
}

# ─────────────────────────────────────────────────
# Constants
HOURS_PER_MONTH = 176.0
N_MONTHS = 6
USD_SCHEDULES = {"schedule b", "b", "schedule d", "d"}
B_D_CATEGORY_FALLBACK = [
    "America (North/South/Canada/Australia)",
    "Middle East/Africa", "Europe", "Asia", "Japan", "Others"
]
AC_CATEGORIES = ["Malaysian", "Regional", "Expatriate"]
# UNIT_TYPES = ["Minimum", "Maximum", "Normalise", "AKER", "DAR", "MMC", "TUAH", "PRW", "PUSB"]
UNIT_TYPES = []
THIRD_PARTY_CATEGORIES = ["Third Party Services"]
NON_LABOUR_CATEGORIES = ["Non-Labour Cost"]

DISCIPLINE_ROW_COUNTS = {
    "General": 9, "Process": 5, "Mechanical Static": 5, "Mechanical Rotating": 5,
    "Mechanical Piping": 5, "Instrument and Control": 5, "Telecommunication": 5,
    "Electrical": 5, "Structural": 5, "Pipeline": 5, "Technical Safety": 5,
    "Material Corrosion Inspection": 5, "HSE": 5
}

DEFAULT_PERSONNEL = {
    "General": ["Project Manager", "Engineering Manager", "Project Engineer",
                "Lead Naval Architecture Engineer", "Planning/Scheduling Engineer",
                "Lead Cost Estimator", "Cost Controller", "General/Clerk Secretary",
                "Document Controller"],
    "Process": ["Lead Engineer Process", "Senior Engineer Process", "Engineer Process", "Drafting", "Designer"],
    "Mechanical Static": ["Lead Engineer Mechanical Static", "Senior Engineer Mechanical Static",
                          "Engineer Mechanical Static", "Drafting", "Designer"],
    "Mechanical Rotating": ["Lead Engineer Mechanical Rotating", "Senior Engineer Mechanical Rotating",
                            "Engineer Mechanical Rotating", "Drafting", "Designer"],
    "Mechanical Piping": ["Lead Engineer Piping", "Senior Engineer Piping", "Engineer Piping", "Drafting", "Designer"],
    "Instrument and Control": ["Lead Engineer Instrument", "Senior Engineer Instrument",
                               "Engineer Instrument", "Drafting", "Designer"],
    "Telecommunication": ["Lead Engineer Telecommunication", "Senior Engineer Telecommunication",
                          "Engineer Telecommunication", "Drafting", "Designer"],
    "Electrical": ["Lead Engineer Electrical", "Senior Engineer Electrical", "Engineer Electrical", "Drafting", "Designer"],
    "Structural": ["Lead Engineer C&S", "Senior Engineer C&S", "Engineer C&S", "Designer", "Drafting"],
    "Pipeline": ["Lead Engineer Pipeline", "Senior Engineer Pipeline", "Engineer Pipeline", "Drafting", "Designer"],
    "Technical Safety": ["Lead Engineer HSE", "Senior Engineer HSE", "Engineer HSE", "Designer", "Drafting"],
    "Material Corrosion Inspection": ["Lead Engineer MCI", "Senior Engineer MCI", "Engineer MCI", "Designer", "Drafting"],
    "HSE": ["Lead Engineer HSE", "Senior Engineer HSE", "Engineer HSE", "Designer", "Drafting"]
}

DISCIPLINE_COLORS = {
    "General": "E6F7F7", "Process": "E8F4FF", "Mechanical Static": "F0F4FF",
    "Mechanical Rotating": "F0FFF4", "Mechanical Piping": "FFF5F5",
    "Instrument and Control": "F5F0FF", "Telecommunication": "E6FFFE",
    "Electrical": "FFF0F8", "Structural": "F5FFF0", "Pipeline": "FFFBEB",
    "Technical Safety": "F0F0FF", "Material Corrosion Inspection": "E6F7F3",
    "HSE": "FFFBF0"
}

DISCIPLINE_SWATCH = {
    "General": "🟦", "Process": "🟩", "Mechanical Static": "🟪",
    "Mechanical Rotating": "🟢", "Mechanical Piping": "🔴",
    "Instrument and Control": "🟫", "Telecommunication": "🔵",
    "Electrical": "🟡", "Structural": "🟠", "Pipeline": "⚫",
    "Technical Safety": "🔶", "Material Corrosion Inspection": "🔷", "HSE": "🟥"
}

# Removed emojis from page labels
PAGES = {
    "MAIN": "Home",
    "TABLE": "Personnel",
    "THIRD_PARTY": "Third Party",
    "NON_LABOUR": "Non-Labour",
    "LOADING": "Loading",
    "TOTALS": "Totals",
    "SUMMARY": "Dashboard",
    "COMPARE": "Compare",
}

GRID_KEY = "grid_df"
THIRD_PARTY_KEY = "third_party_df"
NON_LABOUR_KEY = "non_labour_df"
MONTHLY_LOADING_KEY = "monthly_loading_df"
SAVED_PROJECTS_KEY = "saved_projects"

# ─────────────────────────────────────────────────
# Session state
def init_session_state():
    defaults = {
        "page": "MAIN",
        "dark_mode": False,
        "project_title": "",
        "cost_engineer": "",
        "tp_specialist": "",
        "project_date": None,
        "type_of_package": "U1",
        "type_of_schedule": "Schedule A",
        "rate_source": "MEC.csv",
        "mec_df": pd.DataFrame(),
        "schedule_opts": [],
        "personnel_list": [],
        "rate_types": [],
        "mec_data_loaded": False,
        "mdr_df": pd.DataFrame(),
        "mdr_data_loaded": False,
        GRID_KEY: pd.DataFrame(),
        THIRD_PARTY_KEY: pd.DataFrame(),
        NON_LABOUR_KEY: pd.DataFrame(),
        MONTHLY_LOADING_KEY: pd.DataFrame(),
        SAVED_PROJECTS_KEY: [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_session_state()

# ─────────────────────────────────────────────────
# Global CSS – PETRONAS design + white font for #r28
def inject_css(dark: bool):
    if dark:
        bg          = PETRONAS["navy"]
        bg_alt      = PETRONAS["navy_mid"]
        bg_card     = "#132033"
        sidebar_bg  = PETRONAS["slate"]
        text        = "#E8EFF7"
        text_muted  = PETRONAS["grey_300"]
        border      = "#1E3A5F"
        input_bg    = "#0F2040"
        metric_bg   = "#0F2040"
    else:
        bg          = PETRONAS["off_white"]
        bg_alt      = PETRONAS["grey_100"]
        bg_card     = PETRONAS["white"]
        sidebar_bg  = "#F0F4F8"
        text        = "#0A1628"
        text_muted  = PETRONAS["grey_700"]
        border      = PETRONAS["grey_300"]
        input_bg    = PETRONAS["white"]
        metric_bg   = PETRONAS["white"]

    teal   = PETRONAS["teal"]
    gold   = PETRONAS["gold"]
    teal_d = PETRONAS["teal_dark"]
    teal_l = PETRONAS["teal_light"]

    st.markdown(f"""
    <style>
    /* ── Google Fonts ── */
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=DM+Mono:wght@400;500&display=swap');

    /* ── CSS Variables ── */
    :root {{
        --teal:       {teal};
        --teal-dark:  {teal_d};
        --gold:       {gold};
        --bg:         {bg};
        --bg-alt:     {bg_alt};
        --bg-card:    {bg_card};
        --text:       {text};
        --text-muted: {text_muted};
        --border:     {border};
        --input-bg:   {input_bg};
        --metric-bg:  {metric_bg};
        --radius:     10px;
        --radius-lg:  16px;
        --shadow:     0 4px 24px rgba(0,0,0,0.15);
        --shadow-sm:  0 2px 8px rgba(0,0,0,0.10);
    }}

    /* Force white font for any element with id="r28" */
    span#r28 {{
        color: #FFFFFF !important;
    }}

    /* ── Base ── */
    html, body, .stApp {{
        font-family: 'DM Sans', sans-serif;
        background: var(--bg) !important;
        color: var(--text) !important;
    }}

    /* ── Animated gradient header bar ── */
    .stApp::before {{
        content: '';
        display: block;
        position: fixed;
        top: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, {teal}, {gold}, {teal}, {gold});
        background-size: 300% 100%;
        animation: shimmer 4s ease infinite;
        z-index: 9999;
    }}
    @keyframes shimmer {{
        0%   {{ background-position: 0% 50%; }}
        50%  {{ background-position: 100% 50%; }}
        100% {{ background-position: 0% 50%; }}
    }}

    /* ── Sidebar ── */
    section[data-testid="stSidebar"] {{
        background: {sidebar_bg} !important;
        border-right: 1px solid var(--border);
    }}
    section[data-testid="stSidebar"] * {{
        color: var(--text) !important;
    }}
    section[data-testid="stSidebar"] .stFileUploader label {{
        color: var(--text) !important;
    }}
    /* Sidebar file uploader dropzone — dropzone box always has a light bg, force dark readable text */
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"],
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] * {{
        color: #1a2535 !important;
    }}
    section[data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzoneInstructions"] span,
    section[data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzoneInstructions"] small,
    section[data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzoneInstructions"] p {{
        color: #1a2535 !important;
    }}
    /* Browse files button inside sidebar dropzone */
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button,
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button span {{
        color: #1a2535 !important;
        border-color: #1a2535 !important;
    }}

    /* ── Main content ── */
    .main .block-container {{
        padding: 1.5rem 2.5rem 3rem !important;
        max-width: 1600px;
    }}

    /* ── Typography ── */
    h1, h2, h3, h4 {{
        font-family: 'DM Sans', sans-serif !important;
        font-weight: 600 !important;
        color: var(--text) !important;
        letter-spacing: -0.02em;
    }}
    /* Apply text color to common elements, but NOT inside petronas-banner (those stay white) */
    p:not(.petronas-banner *), label, span:not(.petronas-banner *) {{
        color: var(--text) !important;
    }}
    /* Specifically target standard markdown text blocks rather than global div */
    .stMarkdownContainer {{
        color: var(--text) !important;
    }}
    /* Keep petronas-banner text white unconditionally */
    .petronas-banner h1,
    .petronas-banner h2,
    .petronas-banner h3,
    .petronas-banner p,
    .petronas-banner span {{
        color: #ffffff !important;
    }}
    .petronas-banner .subtitle {{
        color: rgba(255,255,255,0.65) !important;
    }}
    .stCaption, .caption, small {{
        color: var(--text-muted) !important;
        font-size: 0.82rem !important;
    }}
    code, .stCode {{
        font-family: 'DM Mono', monospace !important;
    }}

    /* ── PETRONAS page banner ── */
    .petronas-banner {{
        background: linear-gradient(135deg, {PETRONAS["teal_deep"]} 0%, {PETRONAS["navy_mid"]} 60%, {PETRONAS["slate"]} 100%);
        border-radius: var(--radius-lg);
        padding: 2rem 2.5rem;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
        border: 1px solid rgba(0,161,156,0.3);
        box-shadow: var(--shadow);
    }}
    .petronas-banner::after {{
        content: '';
        position: absolute;
        top: -60px; right: -60px;
        width: 220px; height: 220px;
        border-radius: 50%;
        background: radial-gradient(circle, rgba(0,161,156,0.25) 0%, transparent 70%);
        pointer-events: none;
    }}
    .petronas-banner::before {{
        content: '';
        position: absolute;
        bottom: -40px; left: 30%;
        width: 160px; height: 160px;
        border-radius: 50%;
        background: radial-gradient(circle, rgba(200,169,81,0.12) 0%, transparent 70%);
        pointer-events: none;
    }}
    .petronas-banner h1 {{
        color: #ffffff !important;
        font-size: 2rem !important;
        margin: 0 0 0.25rem 0 !important;
    }}
    .petronas-banner .subtitle {{
        color: rgba(255,255,255,0.65) !important;
        font-size: 0.92rem;
        margin: 0;
    }}
    .petronas-banner .badge {{
        display: inline-flex;
        align-items: center;
        gap: 0.4rem;
        background: rgba(0,161,156,0.25);
        border: 1px solid rgba(0,161,156,0.5);
        color: #7FD8D5 !important;
        padding: 0.3rem 0.9rem;
        border-radius: 20px;
        font-size: 0.78rem;
        font-weight: 600;
        letter-spacing: 0.04em;
        text-transform: uppercase;
        margin-top: 1rem;
    }}

    /* ── Cards ── */
    .pcard {{
        background: var(--bg-card);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: 1.4rem 1.6rem;
        margin-bottom: 1rem;
        box-shadow: var(--shadow-sm);
        transition: box-shadow 0.2s, border-color 0.2s;
    }}
    .pcard:hover {{
        border-color: var(--teal);
        box-shadow: 0 0 0 2px rgba(0,161,156,0.12), var(--shadow-sm);
    }}
    .pcard-accent {{
        border-left: 3px solid var(--teal);
    }}
    .pcard-gold {{
        border-left: 3px solid var(--gold);
    }}

    /* ── Metric cards ── */
    .metric-card {{
        background: var(--metric-bg);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        padding: 1.25rem 1.5rem;
        margin-bottom: 0.75rem;
        position: relative;
        overflow: hidden;
        transition: transform 0.2s, box-shadow 0.2s;
    }}
    .metric-card:hover {{
        transform: translateY(-2px);
        box-shadow: 0 8px 24px rgba(0,161,156,0.15);
    }}
    .metric-card::before {{
        content: '';
        position: absolute;
        top: 0; left: 0;
        width: 100%; height: 3px;
        background: linear-gradient(90deg, var(--teal), var(--gold));
    }}
    .metric-card .label {{
        color: var(--text-muted) !important;
        font-size: 0.78rem;
        font-weight: 600;
        letter-spacing: 0.06em;
        text-transform: uppercase;
        margin-bottom: 0.5rem;
    }}
    .metric-card .value {{
        color: var(--text) !important;
        font-size: 1.75rem;
        font-weight: 700;
        line-height: 1;
        font-variant-numeric: tabular-nums;
    }}
    .metric-card .sub {{
        color: var(--text-muted) !important;
        font-size: 0.8rem;
        margin-top: 0.35rem;
    }}

    /* ── Summary card ── */
    .summary-card {{
        background: var(--bg-card);
        border: 1px solid var(--border);
        border-radius: var(--radius-lg);
        padding: 1.75rem 2rem;
        margin-bottom: 1.5rem;
        box-shadow: var(--shadow-sm);
    }}

    /* ── Divider ── */
    .pdivider {{
        height: 1px;
        background: linear-gradient(90deg, var(--teal) 0%, transparent 100%);
        opacity: 0.3;
        margin: 1.25rem 0;
        border: none;
    }}

    /* ── Section header ── */
    .section-header {{
        display: flex;
        align-items: center;
        gap: 0.6rem;
        margin-bottom: 1rem;
        padding-bottom: 0.6rem;
        border-bottom: 1px solid var(--border);
    }}
    .section-header .dot {{
        width: 8px; height: 8px;
        border-radius: 50%;
        background: var(--teal);
        flex-shrink: 0;
    }}
    .section-header h3 {{
        margin: 0 !important;
        font-size: 1rem !important;
        font-weight: 600 !important;
        color: var(--text) !important;
    }}

    /* ── Nav tabs ── */
    /* Target the nav layout specifically */
    div.element-container:has(.nav-marker) + div[data-testid="stHorizontalBlock"] {{
        flex-wrap: wrap !important;
    }}
    div.element-container:has(.nav-marker) + div[data-testid="stHorizontalBlock"] > div[data-testid="column"] {{
        min-width: 90px !important;
        flex: 1 1 auto !important;
    }}

    /* ── Streamlit buttons override ── */
    div.stButton > button {{
        font-family: 'DM Sans', sans-serif !important;
        font-weight: 500 !important;
        border-radius: 8px !important;
        transition: all 0.18s !important;
        letter-spacing: 0.01em;
        white-space: nowrap !important;
        padding-left: 0.25rem !important;
        padding-right: 0.25rem !important;
        font-size: 0.8rem !important;
        min-height: 2.2rem !important;
    }}
    div.stButton > button[kind="primary"] {{
        background: linear-gradient(135deg, {teal} 0%, {teal_d} 100%) !important;
        border: none !important;
        color: white !important;
        box-shadow: 0 2px 8px rgba(0,161,156,0.35) !important;
    }}
    div.stButton > button[kind="primary"]:hover {{
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 16px rgba(0,161,156,0.45) !important;
    }}
    div.stButton > button[kind="secondary"] {{
        background: transparent !important;
        border: 1px solid var(--border) !important;
        color: var(--text) !important;
    }}
    div.stButton > button[kind="secondary"]:hover {{
        border-color: var(--teal) !important;
        color: var(--teal) !important;
        background: rgba(0,161,156,0.06) !important;
    }}

    /* ── Inputs ── */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div,
    .stSelectbox div[data-baseweb="select"] span,
    .stDateInput > div > div > input {{
        background: var(--input-bg) !important;
        border-color: var(--border) !important;
        border-radius: 8px !important;
        color: var(--text) !important;
        font-family: 'DM Sans', sans-serif !important;
    }}
    
    /* Fix for dropdown menus in dark mode */
    ul[role="listbox"], ul[role="listbox"] li {{
        background: var(--bg-card) !important;
        color: var(--text) !important;
    }}
    div[data-baseweb="popover"] {{
        background: var(--bg-card) !important;
    }}
    .stSelectbox div[data-baseweb="select"] [data-testid="stMarkdownContainer"] p {{
        color: var(--text) !important;
    }}
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {{
        border-color: var(--teal) !important;
        box-shadow: 0 0 0 2px rgba(0,161,156,0.15) !important;
    }}
    .stSelectbox > div > div:focus-within {{
        border-color: var(--teal) !important;
        box-shadow: 0 0 0 2px rgba(0,161,156,0.15) !important;
    }}
    label[data-testid="stWidgetLabel"] p {{
        color: var(--text-muted) !important;
        font-size: 0.82rem !important;
        font-weight: 500 !important;
        letter-spacing: 0.02em;
        text-transform: uppercase;
    }}

    /* ── Info / warning / error boxes ── */
    .stAlert {{
        border-radius: var(--radius) !important;
        border-left-width: 3px !important;
        font-size: 0.9rem !important;
    }}
    div[data-testid="stInfo"] {{
        background: rgba(0,161,156,0.08) !important;
        border-left-color: var(--teal) !important;
        color: var(--text) !important;
    }}
    div[data-testid="stWarning"] {{
        background: rgba(245,158,11,0.08) !important;
        border-left-color: {PETRONAS["warning"]} !important;
    }}
    div[data-testid="stSuccess"] {{
        background: rgba(34,197,94,0.08) !important;
        border-left-color: {PETRONAS["success"]} !important;
    }}
    div[data-testid="stError"] {{
        background: rgba(239,68,68,0.08) !important;
        border-left-color: {PETRONAS["danger"]} !important;
    }}

    /* ── Streamlit metrics ── */
    div[data-testid="metric-container"] {{
        background: var(--metric-bg) !important;
        border: 1px solid var(--border) !important;
        border-radius: var(--radius) !important;
        padding: 1rem 1.25rem !important;
    }}
    div[data-testid="metric-container"] label p {{
        color: var(--text-muted) !important;
        font-size: 0.78rem !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        font-weight: 600 !important;
    }}
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {{
        color: var(--text) !important;
        font-weight: 700 !important;
    }}

    /* ── Dataframe ── */
    .stDataFrame {{
        border: 1px solid var(--border) !important;
        border-radius: var(--radius) !important;
        overflow: hidden;
    }}
    .stDataFrame thead tr th {{
        background: var(--bg-alt) !important;
        color: var(--text-muted) !important;
        font-size: 0.78rem !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }}
    /* Handle text colors for HTML-based dataframes */
    .stDataFrame table, .stDataFrame td, .stDataFrame th {{
        color: var(--text) !important;
        background: var(--bg-card);
    }}
    .stDataFrame div[data-testid="stDataFrame"] * {{
        color: var(--text) !important;
    }}

    /* ── Comparison badges ── */
    .comparison-badge {{
        display: inline-flex;
        align-items: center;
        gap: 0.25rem;
        padding: 0.2rem 0.7rem;
        border-radius: 20px;
        font-size: 0.78rem;
        font-weight: 700;
        letter-spacing: 0.02em;
    }}
    .comparison-badge.up {{
        background: rgba(34,197,94,0.12);
        color: #22C55E !important;
        border: 1px solid rgba(34,197,94,0.3);
    }}
    .comparison-badge.down {{
        background: rgba(239,68,68,0.12);
        color: #EF4444 !important;
        border: 1px solid rgba(239,68,68,0.3);
    }}
    .comparison-badge.neutral {{
        background: rgba(122,142,163,0.12);
        color: {PETRONAS["grey_500"]} !important;
        border: 1px solid rgba(122,142,163,0.3);
    }}

    /* ── Sidebar logo area ── */
    .sidebar-logo {{
        display: flex;
        align-items: center;
        gap: 0.75rem;
        padding: 1rem 0 0.5rem 0;
        margin-bottom: 0.75rem;
        border-bottom: 1px solid var(--border);
    }}
    .sidebar-logo .logo-icon {{
        width: 44px; height: 44px;
        border-radius: 8px;
        display: flex;
        align-items: center;
        justify-content: center;
        flex-shrink: 0;
        overflow: hidden;
    }}
    .sidebar-logo .logo-icon img {{
        width: 44px; height: 44px;
        object-fit: contain;
        border-radius: 8px;
    }}
    .sidebar-logo .logo-text {{
        font-weight: 700;
        font-size: 1rem;
        color: var(--text) !important;
        line-height: 1.1;
    }}
    .sidebar-logo .logo-sub {{
        font-size: 0.7rem;
        color: var(--text-muted) !important;
        letter-spacing: 0.04em;
    }}

    /* ── Step indicator ── */
    .step-bar {{
        display: flex;
        gap: 0.5rem;
        margin-bottom: 1.5rem;
        align-items: center;
        flex-wrap: wrap;
    }}
    .step-item {{
        display: flex;
        align-items: center;
        gap: 0.4rem;
        font-size: 0.78rem;
        font-weight: 500;
        color: var(--text-muted) !important;
    }}
    .step-item.active {{
        color: var(--teal) !important;
        font-weight: 700;
    }}
    .step-dot {{
        width: 22px; height: 22px;
        border-radius: 50%;
        background: var(--bg-alt);
        border: 1.5px solid var(--border);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 0.68rem;
        font-weight: 700;
        flex-shrink: 0;
    }}
    .step-dot.active {{
        background: var(--teal);
        border-color: var(--teal);
        color: white !important;
        box-shadow: 0 0 0 4px rgba(0,161,156,0.2);
    }}
    .step-dot.done {{
        background: rgba(0,161,156,0.15);
        border-color: var(--teal);
        color: var(--teal) !important;
    }}
    .step-sep {{
        flex: 1;
        min-width: 16px;
        height: 1px;
        background: var(--border);
    }}

    /* ── Loading spinner ── */
    .stSpinner > div {{
        border-top-color: var(--teal) !important;
    }}

    /* ── Radio / toggle ── */
    .stRadio label {{
        font-size: 0.88rem !important;
    }}
    .stToggle label {{
        font-size: 0.88rem !important;
    }}

    /* ── Hide streamlit branding ── */
    #MainMenu {{ display: none !important; }}
    footer {{ display: none !important; }}
    .stAppDeployButton {{ display: none !important; }}
    
    /* Make header transparent but keep toggles visible */
    header[data-testid="stHeader"] {{
        background: transparent !important;
        border-bottom: none !important;
    }}

    /* Sidebar toggle styling */
    [data-testid="collapsedControl"] {{
        z-index: 99999 !important;
        background: var(--teal) !important;
        border-radius: 6px !important;
        margin: 0.5rem;
    }}
    [data-testid="collapsedControl"] svg {{
        fill: #ffffff !important;
    }}
    
    .main .block-container {{
        padding-top: 1.5rem !important;
    }}

    /* ── File uploader ── */
    .stFileUploader > div {{
        border: 1.5px dashed var(--border) !important;
        border-radius: var(--radius) !important;
        background: var(--bg-alt) !important;
        transition: border-color 0.2s;
    }}
    .stFileUploader > div:hover {{
        border-color: var(--teal) !important;
    }}
    /* Fix drag-and-drop text always visible in both modes */
    .stFileUploader small, .stFileUploader span, .stFileUploader p,
    [data-testid="stFileUploaderDropzone"] small,
    [data-testid="stFileUploaderDropzone"] span,
    [data-testid="stFileUploaderDropzone"] p {{
        color: var(--text-muted) !important;
    }}
    [data-testid="stFileUploaderDropzoneInstructions"] > div > span {{
        color: var(--text) !important;
    }}

    /* ── Scrollbar ── */
    ::-webkit-scrollbar {{ width: 6px; height: 6px; }}
    ::-webkit-scrollbar-track {{ background: var(--bg-alt); }}
    ::-webkit-scrollbar-thumb {{
        background: var(--teal);
        border-radius: 3px;
        opacity: 0.5;
    }}
    ::-webkit-scrollbar-thumb:hover {{ background: var(--teal-dark); }}

    /* ── Fade-in animation ── */
    @keyframes fadeUp {{
        from {{ opacity: 0; transform: translateY(12px); }}
        to   {{ opacity: 1; transform: translateY(0); }}
    }}
    .fadeup {{
        animation: fadeUp 0.35s ease forwards;
    }}

    /* ── Pulse for status badge ── */
    @keyframes pulse {{
        0%, 100% {{ opacity: 1; }}
        50%       {{ opacity: 0.5; }}
    }}
    .pulse {{ animation: pulse 2s ease-in-out infinite; }}

    </style>
    """, unsafe_allow_html=True)

inject_css(st.session_state["dark_mode"])

# Load logo once — used in sidebar and app header
_logo_b64  = base64.b64encode(_logo_path.read_bytes()).decode() if _logo_path.exists() else ""
_logo_html = (f'<img src="data:image/png;base64,{_logo_b64}" '
              f'style="width:44px;height:44px;object-fit:contain;border-radius:8px">'
             ) if _logo_b64 else "⛽"

# ─────────────────────────────────────────────────
# Sidebar
with st.sidebar:
    # Logo
    st.markdown(f"""
    <div class="sidebar-logo">
        <div class="logo-icon">{_logo_html}</div>
        <div>
            <div class="logo-text">MEC TOOL</div>
            <div class="logo-sub">PETRONAS UPSTREAM CE</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    dark_mode = st.toggle("Dark Mode", value=st.session_state["dark_mode"])
    if dark_mode != st.session_state["dark_mode"]:
        st.session_state["dark_mode"] = dark_mode
        st.rerun()

    st.markdown("<hr style='border:none;border-top:1px solid var(--border);margin:0.75rem 0'>", unsafe_allow_html=True)

    st.markdown("**MEC Rates**")
    mec_file = st.file_uploader("Upload MEC.csv", type=["csv"], key="mec_uploader", label_visibility="collapsed")
    if mec_file is not None:
        st.success("✅ MEC.csv ready")
        if st.button("Load MEC Data", use_container_width=True, type="primary"):
            st.session_state["mec_data_loaded"] = False
            st.cache_data.clear()
            st.rerun()
    else:
        st.warning("⬆️ Upload MEC.csv to begin")

    st.markdown("<hr style='border:none;border-top:1px solid var(--border);margin:0.75rem 0'>", unsafe_allow_html=True)

    st.markdown("**MDR Hours**")
    mdr_file = st.file_uploader("Upload MDR.csv", type=["csv"], key="mdr_uploader", label_visibility="collapsed")
    if mdr_file is not None:
        _err = st.session_state.get("mdr_load_error")
        if _err:
            st.error(f"**MDR Load Error:** {_err}")
        else:
            st.success("✅ MDR.csv ready")
            
        if st.button("Load MDR Data", use_container_width=True, type="primary"):
            st.session_state["mdr_data_loaded"] = False
            st.session_state["mdr_load_error"] = None
            st.rerun()
    else:
        st.info("MDR.csv optional")

    st.markdown("<hr style='border:none;border-top:1px solid var(--border);margin:0.75rem 0'>", unsafe_allow_html=True)

    # ── MDR Debug ──────────────────────────────────────────────────────
    with st.expander("🔍 MDR Debug", expanded=False):
        st.write(f"**mdr_file:** `{'None' if mdr_file is None else mdr_file.name}`")
        st.write(f"**mdr_data_loaded:** `{st.session_state.get('mdr_data_loaded', '?')}`")
        _dbg_mdr = st.session_state.get("mdr_df", None)
        st.write(f"**session mdr_df type:** `{type(_dbg_mdr).__name__}`")
        if isinstance(_dbg_mdr, pd.DataFrame):
            st.write(f"**shape:** `{_dbg_mdr.shape}`, **columns:** `{list(_dbg_mdr.columns)}`")
            if not _dbg_mdr.empty:
                st.dataframe(_dbg_mdr.head(5))
        elif isinstance(_dbg_mdr, tuple):
            st.warning(f"⚠️ mdr_df is a tuple — stale value: {type(_dbg_mdr[0]).__name__}, err={_dbg_mdr[1] if len(_dbg_mdr)>1 else 'n/a'}")
        else:
            st.write("mdr_df not set yet")

        if mdr_file is not None:
            mdr_file.seek(0)
            _raw = mdr_file.read()
            st.write(f"**File size:** {len(_raw)} bytes")
            st.write("**First 10 raw lines:**")
            st.code(b'\n'.join(_raw.splitlines()[:10]).decode('utf-8', errors='replace'))
            st.info("Click **Load MDR Data** above to run conversion and apply to grid.")
    # ── end MDR Debug ──────────────────────────────────────────────────

    # Live status
    sched_now = st.session_state.get("type_of_schedule", "—")
    pkg_now   = st.session_state.get("type_of_package", "—")
    title_now = st.session_state.get("project_title", "—") or "Untitled"
    st.markdown(f"""
    <div style="font-size:0.78rem; color:var(--text-muted); line-height:1.8">
        <div>📋 Schedule: <strong style="color:var(--teal)">{sched_now}</strong></div>
        <div>📦 Package: <strong style="color:var(--teal)">{pkg_now}</strong></div>
    </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# Helpers (unchanged)
NBSP = "\xa0"
_num = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?")

def _to_float_safe(val):
    try:
        if val is None: return 0.0
        if isinstance(val, (int, float)):
            return 0.0 if pd.isna(val) else float(val)
        s = re.sub(r"[$,€£\s]", "", str(val).strip())
        s = re.sub(r"[^\d.-]", " ", s)
        m = _num.search(s)
        return float(m.group()) if m else 0.0
    except Exception:
        return 0.0

def _canon(s):
    return re.sub(r"\s+", " ", str(s or "").replace(NBSP, " ").strip().lower())

def _canon_sched_tag(s):
    t = _canon(s)
    m = re.search(r"(?:schedule\s*)?([abcd])$", t)
    return m.group(1) if m else t

def get_col(df, name):
    if df.empty: return pd.Series([])
    if name in df.columns: return df[name]
    name_lower = name.lower()
    for col in df.columns:
        if col.lower() == name_lower or name_lower in col.lower():
            return df[col]
    return pd.Series([None] * len(df))

def is_usd(schedule):
    return _canon(schedule) in USD_SCHEDULES or _canon_sched_tag(schedule) in {"b", "d"}

def currency_for(schedule):
    return "USD" if is_usd(schedule) else "MYR"

def build_category_options(schedule, mec_df):
    if mec_df is not None and not mec_df.empty:
        df = mec_df.copy()
        if "SCHEDULES" in df.columns and schedule:
            sched_col = get_col(df, "SCHEDULES")
            if not sched_col.empty:
                df = df[sched_col.astype(str).str.contains(schedule, case=False, na=False)]
        if "CATEGORY" in df.columns:
            cats = get_col(df, "CATEGORY").dropna().astype(str).str.strip().unique().tolist()
            cats = [c for c in cats if c and c.lower() != 'nan']
            if cats:
                return sorted(cats)
    return B_D_CATEGORY_FALLBACK if is_usd(schedule) else AC_CATEGORIES

# ─────────────────────────────────────────────────
# Grid init
def initialize_default_grids():
    tmp_schedule = st.session_state["type_of_schedule"]
    categories = build_category_options(tmp_schedule, st.session_state.get("mec_df", pd.DataFrame()))
    if not categories:
        categories = B_D_CATEGORY_FALLBACK if is_usd(tmp_schedule) else AC_CATEGORIES

    _mdr_raw = st.session_state.get("mdr_df", pd.DataFrame())
    mdr_df = _mdr_raw[0] if isinstance(_mdr_raw, tuple) else _mdr_raw
    
    num_months = N_MONTHS
    if not mdr_df.empty:
        month_cols = [c for c in mdr_df.columns if str(c).startswith('Month ')]
        if month_cols:
            num_months = len(month_cols)
            
    rows = []
    if not mdr_df.empty and all(c in mdr_df.columns for c in ["DISCIPLINE", "PERSONNEL", "Total (Hours)"]):
        for _, row in mdr_df.iterrows():
            disc = str(row["DISCIPLINE"]).strip()
            pers = str(row["PERSONNEL"]).strip()
            try: total_hrs = float(row["Total (Hours)"])
            except: total_hrs = 0.0
            weightage = total_hrs / HOURS_PER_MONTH
            rows.append({
                "Swatch": DISCIPLINE_SWATCH.get(disc, "⬜"),
                "Discipline": disc, "Personnel": pers,
                "Category": categories[0] if categories else "",
                "Type of Unit Rate": "Normalise",
                "Total Hours": total_hrs,
                "Weightage (FTE)": weightage,
                "Weightage / Month": weightage / num_months if num_months else 0.0
            })
    else:
        for disc, count in DISCIPLINE_ROW_COUNTS.items():
            defaults = DEFAULT_PERSONNEL.get(disc, [])
            for i in range(count):
                pers = defaults[i] if i < len(defaults) else ""
                rows.append({
                    "Swatch": DISCIPLINE_SWATCH.get(disc, "⬜"),
                    "Discipline": disc, "Personnel": pers,
                    "Category": categories[0] if categories else "",
                    "Type of Unit Rate": "Normalise",
                    "Total Hours": 0.0, "Weightage (FTE)": 0.0,
                    "Weightage / Month": 0.0
                })
    st.session_state[GRID_KEY] = pd.DataFrame(rows)

def reset_grid():
    initialize_default_grids()

def reset_all():
    st.session_state["project_title"] = ""
    st.session_state["cost_engineer"] = ""
    st.session_state["tp_specialist"] = ""
    st.session_state["project_date"] = None
    st.session_state["type_of_package"] = "U1"
    if st.session_state.get("schedule_opts"):
        st.session_state["type_of_schedule"] = st.session_state["schedule_opts"][0]
    else:
        st.session_state["type_of_schedule"] = "Schedule A"
    initialize_default_grids()
    months = [f"Month {i+1:02d}" for i in range(12)]
    st.session_state[THIRD_PARTY_KEY] = pd.DataFrame([{"Category": "Third Party Services", "Description": "",
                                                        "Basis": "Percentage of Labour Cost", "Percentage": 0.0,
                                                        "Fixed Amount": 0.0, "Remarks": ""}])
    st.session_state[NON_LABOUR_KEY] = pd.DataFrame([{"Category": "Non-Labour Cost", "Description": "",
                                                       "Basis": "Percentage of Labour Cost", "Percentage": 0.0,
                                                       "Fixed Amount": 0.0, "Remarks": ""}])
    st.session_state[MONTHLY_LOADING_KEY] = pd.DataFrame({
        "Month": months,
        "Loading Factor (%)": [100.0]*12,
        "Weightage Distribution": [100.0/12]*12
    })
    st.session_state["page"] = "MAIN"
    st.success("Project reset. Saved projects preserved.")

# ─────────────────────────────────────────────────
# CSV loaders
@st.cache_data(show_spinner=False, ttl=600)
def load_mec_csv(file_obj):
    try:
        file_obj.seek(0)
        df = pd.read_csv(file_obj)
        df.columns = [str(col).strip() for col in df.columns]
        rename = {}
        for col in df.columns:
            low = col.lower().strip()
            if low == 'project': rename[col] = 'PROJECT'
            elif low == 'project description': rename[col] = 'PROJECT DESCRIPTION'
            elif low == 'schedules': rename[col] = 'SCHEDULES'
            elif low == 'schedule description': rename[col] = 'SCHEDULES DESCRIPTION'
            elif low == 'category': rename[col] = 'CATEGORY'
            elif low == 'personnel': rename[col] = 'PERSONNEL'
            elif low == 'type of rate': rename[col] = 'TYPE OF RATE'
            elif low == 'unit rate': rename[col] = 'UNIT RATE'
        df = df.rename(columns=rename)
        for req in ['PROJECT', 'PROJECT DESCRIPTION', 'SCHEDULES', 'SCHEDULES DESCRIPTION',
                    'CATEGORY', 'PERSONNEL', 'TYPE OF RATE', 'UNIT RATE']:
            if req not in df.columns: df[req] = None
        return df
    except Exception as e:
        st.warning(f"Error reading MEC.csv: {e}")
        return pd.DataFrame()

def load_mdr_csv(file_obj):
    import io
    import numpy as np
    try:
        file_obj.seek(0)
        raw_bytes = file_obj.read()
        df_raw = pd.read_csv(io.BytesIO(raw_bytes), header=None, dtype=str, keep_default_na=False)

        # Helper to strip strings
        stripper = np.vectorize(lambda x: x.strip() if isinstance(x, str) else x)
        df_raw = df_raw.apply(stripper)

        # 1) Try to detect hours/month from the preheader row that contains "Hours" followed by repeating numeric values (e.g., 176)
        hours_per_month = 176.0
        hours_row_idx = None
        for i in range(len(df_raw)):
            row = df_raw.iloc[i].tolist()
            row_lower = [str(cell).lower() for cell in row]
            if 'hours' in row_lower:
                j = row_lower.index('hours')
                nums = []
                for cell in row[j+1:]:
                    s = str(cell).replace(',', '')
                    try:
                        val = float(s)
                        nums.append(val)
                    except:
                        pass
                if nums:
                    hours_per_month = nums[0]
                    hours_row_idx = i
                    break

        # 2) Find the header row that starts the main table (contains DISCIPLINE, PERSONNEL)
        header_idx = None
        for i in range(len(df_raw)):
            row = [str(x).upper().strip() for x in df_raw.iloc[i].tolist()]
            if any('DISCIPLINE' in c for c in row) and any('PERSONNEL' in c.replace(' ', '') for c in row):
                header_idx = i
                break

        if header_idx is None:
            # Fallback for pre-formatted CSV
            df = pd.read_csv(io.BytesIO(raw_bytes))
            df.columns = [str(col).strip() for col in df.columns]
            missing = [c for c in ["DISCIPLINE", "PERSONNEL", "Total (Hours)"] if c not in df.columns]
            if missing:
                return pd.DataFrame(), f"Could not find table header row containing DISCIPLINE, PERSONNEL. Also missing expected preformatted columns: {missing}"
            df = df[["DISCIPLINE", "PERSONNEL", "Total (Hours)"]].copy()
            df["Total (Hours)"] = pd.to_numeric(df["Total (Hours)"], errors="coerce").fillna(0)
            df = df.dropna(subset=["DISCIPLINE", "PERSONNEL"])
            df["DISCIPLINE"] = df["DISCIPLINE"].astype(str).str.strip()
            df["PERSONNEL"]  = df["PERSONNEL"].astype(str).str.strip()
            return df, None

        header = df_raw.iloc[header_idx].tolist()
        # Build a DataFrame of the data rows under the header
        body = df_raw.iloc[header_idx+1:].copy()
        body.columns = header + [f'_extra_{k}' for k in range(max(0, body.shape[1] - len(header)))]

        # Normalize column names (strip extra spaces)
        body.columns = [str(c).strip() for c in body.columns]

        # 3) Locate key columns by name
        disc_col = next((c for c in body.columns if c.strip().upper() == 'DISCIPLINE'), None)
        pers_col = next((c for c in body.columns if c.strip().upper() in ('PERSONNEL','PERSONNEL,','PERSONNEL;','PERSONNEL:')), None)
        pers_idx = next((i for i, c in enumerate(body.columns) if c.strip().upper() in ('PERSONNEL','PERSONNEL,','PERSONNEL;','PERSONNEL:')), None)
        rate_col = next((i for i, c in enumerate(body.columns) if str(c).strip().upper().startswith('UNIT RATE')), None)

        if disc_col is None or pers_col is None:
            diag = {
                'disc_col': disc_col,
                'pers_col': pers_col,
                'rate_col_idx': rate_col,
                'columns_sample': body.columns[:30].tolist()
            }
            return pd.DataFrame(), f'Failed to detect key columns (DISCIPLINE/PERSONNEL). Diagnostics: {diag}'

        # 4) Determine month columns dynamically.
        start_idx = (rate_col + 1) if rate_col is not None else (pers_idx + 1)
        raw_headers_upper = [str(c).strip().upper() for c in body.columns]

        num_months = None

        if hours_row_idx is not None:
            hours_row = df_raw.iloc[hours_row_idx].tolist()
            count = 0
            for pos in range(start_idx, len(hours_row)):
                val = str(hours_row[pos]).replace(',', '').strip()
                try:
                    if float(val) > 0:
                        count += 1
                    else:
                        break           # stop at the first zero / blank
                except ValueError:
                    break
            if count > 0:
                num_months = count

        if num_months is None:
            # Fallback: scan headers for "TOTAL" or "HOUR" sentinel
            end_idx = len(body.columns)
            for pos in range(start_idx, len(body.columns)):
                h = raw_headers_upper[pos]
                if 'TOTAL' in h or ('HOUR' in h and pos > start_idx):
                    end_idx = pos
                    break
            num_months = end_idx - start_idx

        end_idx = start_idx + num_months

        if num_months < 1:
            return pd.DataFrame(), f'Could not detect any month columns after unit-rate column (index {rate_col}). Headers from that position: {raw_headers_upper[start_idx:start_idx+15]}'

        new_cols = list(body.columns)
        for i, pos in enumerate(range(start_idx, end_idx)):
            new_cols[pos] = f'_m{i + 1}'
        body.columns = new_cols
        month_cols = [f'_m{i + 1}' for i in range(num_months)]

        # 5) Clean and filter rows: drop subtotal rows and empty personnel
        def is_total_row(x):
            return isinstance(x, str) and x.strip().upper().startswith('TOTAL PER DISCIPLINE')

        mask_valid = (~body[disc_col].apply(is_total_row)) & (body[pers_col].astype(str).str.len() > 0)
        body_valid = body.loc[mask_valid].copy()

        # 6) Coerce month columns to numeric multipliers
        for c in month_cols:
            body_valid[c] = (body_valid[c].astype(str)
                             .str.replace(',', '', regex=False)
                             .str.replace('%', '', regex=False)
                             .replace({'': np.nan}))
            body_valid[c] = pd.to_numeric(body_valid[c], errors='coerce').fillna(0.0)

        # 7) Compute Total Hours from months * hours_per_month (float), even if a total hours column exists
        month_sum = body_valid[month_cols].sum(axis=1)
        body_valid['Total (Hours)'] = month_sum * hours_per_month

        # 8) Build MDR DataFrame
        mdr_df = pd.DataFrame({
            'DISCIPLINE': body_valid[disc_col].astype(str).str.strip(),
            'PERSONNEL': body_valid[pers_col].astype(str).str.strip(),
        })
        for i, c in enumerate(month_cols, start=1):
            mdr_df[f'Month {i}'] = body_valid[c].values
        mdr_df['Total (Hours)'] = body_valid['Total (Hours)'].values

        # 9) Remove fully empty rows (if any) and reset index
        is_all_zero = (mdr_df[[f'Month {i}' for i in range(1, len(month_cols)+1)]].sum(axis=1) == 0)
        mdr_df = mdr_df.loc[~(mdr_df['PERSONNEL'].eq('') & is_all_zero)].reset_index(drop=True)

        return mdr_df, None

    except Exception as e:
        import traceback
        return pd.DataFrame(), f"Error converting MDR: {e}\n{traceback.format_exc()}"

# ─────────────────────────────────────────────────
# Load data
if mec_file is not None and not st.session_state["mec_data_loaded"]:
    with st.spinner("Loading MEC.csv…"):
        mec_df = load_mec_csv(mec_file)
        if mec_df.empty:
            st.error("Failed to load MEC.csv")
            st.stop()
        schedule_opts = []
        if "SCHEDULES" in mec_df.columns:
            s = get_col(mec_df, "SCHEDULES").dropna().astype(str).str.strip()
            schedule_opts = sorted([v for v in s.unique() if v and v.lower() != 'nan'])
        personnel_list = []
        if "PERSONNEL" in mec_df.columns:
            personnel_list.extend(get_col(mec_df, "PERSONNEL").dropna().astype(str).tolist())
        for plist in DEFAULT_PERSONNEL.values():
            personnel_list.extend(plist)
        personnel_list = sorted(set(p for p in personnel_list if p and p.lower() != 'nan'))
        rate_types = []
        if "TYPE OF RATE" in mec_df.columns:
            rate_types.extend(get_col(mec_df, "TYPE OF RATE").dropna().astype(str).str.strip().unique().tolist())
        for r in UNIT_TYPES:
            if r not in rate_types: rate_types.append(r)
        rate_types = sorted(set(rate_types))
        st.session_state["mec_df"] = mec_df
        st.session_state["schedule_opts"] = schedule_opts
        st.session_state["personnel_list"] = personnel_list
        st.session_state["rate_types"] = rate_types
        st.session_state["mec_data_loaded"] = True
        if schedule_opts:
            st.session_state["type_of_schedule"] = schedule_opts[0]
        initialize_default_grids()
        st.rerun()

if mdr_file is not None and not st.session_state["mdr_data_loaded"]:
    with st.spinner("Loading MDR.csv…"):
        _mdr_result = load_mdr_csv(mdr_file)
        if isinstance(_mdr_result, tuple):
            _mdr_loaded, _mdr_err = _mdr_result
        else:
            _mdr_loaded, _mdr_err = _mdr_result, None
            
        if _mdr_err:
            st.session_state["mdr_load_error"] = _mdr_err
        else:
            st.session_state["mdr_load_error"] = None
            st.success(f"MDR loaded: {len(_mdr_loaded)} rows, columns: {list(_mdr_loaded.columns)}")
            
        st.session_state["mdr_df"] = _mdr_loaded
        st.session_state["mdr_data_loaded"] = True
        initialize_default_grids()
        st.rerun()

mec_df         = st.session_state.get("mec_df", pd.DataFrame())
schedule_opts  = st.session_state.get("schedule_opts", [])
PERSONNEL_LIST = st.session_state.get("personnel_list", [])
RATE_TYPES     = st.session_state.get("rate_types", UNIT_TYPES)
_mdr_raw       = st.session_state.get("mdr_df", pd.DataFrame())
mdr_df         = _mdr_raw[0] if isinstance(_mdr_raw, tuple) else _mdr_raw

if st.session_state["mec_data_loaded"] and st.session_state[GRID_KEY].empty:
    initialize_default_grids()

# ─────────────────────────────────────────────────
# Rate lookup
def _canonical_col(df, name):
    s = get_col(df, name)
    return s.astype(str).str.replace(NBSP, " ").str.strip().str.lower() if not s.empty else pd.Series([])

def _relaxed_match(df, personnel, category, schedule, rate_type, package):
    if df.empty: return df
    m = df.copy()
    if "PROJECT" in m.columns and package:
        pkg_s = _canonical_col(m, "PROJECT")
        if not pkg_s.empty:
            m = m[pkg_s.str.contains(package.lower().strip(), na=False)]
    if "PERSONNEL" in m.columns and personnel:
        pers_s = _canonical_col(m, "PERSONNEL")
        if not pers_s.empty:
            pt = personnel.lower().strip()
            exact = pers_s == pt
            m = m[exact] if exact.any() else m[pers_s.str.contains(pt, na=False)]
    if "CATEGORY" in m.columns and category:
        cat_s = _canonical_col(m, "CATEGORY")
        if not cat_s.empty:
            ct = category.lower().strip()
            exact = cat_s == ct
            m = m[exact] if exact.any() else m[cat_s.str.contains(ct, na=False)]
    if schedule and "SCHEDULES" in m.columns:
        sch_s = _canonical_col(m, "SCHEDULES")
        if not sch_s.empty:
            m = m[sch_s.str.contains(_canon_sched_tag(schedule), na=False)]
    if "TYPE OF RATE" in m.columns and rate_type:
        rate_s = _canonical_col(m, "TYPE OF RATE")
        if not rate_s.empty:
            rt = rate_type.lower().strip()
            exact = rate_s == rt
            m = m[exact] if exact.any() else m[rate_s.str.contains(rt, na=False)]
    return m

def get_rate(mec_df, personnel, category, unit_type, schedule, package):
    if mec_df.empty: return 0.0
    try:
        matches = _relaxed_match(mec_df, personnel, category, schedule, unit_type, package)
        if matches.empty or "UNIT RATE" not in matches.columns: return 0.0
        return _to_float_safe(matches["UNIT RATE"].iat[0])
    except Exception:
        return 0.0

# ─────────────────────────────────────────────────
# Calculations
def calculate_labour_costs(grid_df, currency, type_of_schedule, type_of_package):
    out_cols = ["Discipline", "Personnel", "Category", "Type of Unit Rate",
                f"Unit Rate ({currency})", "Total Hours", "Weightage (FTE)",
                "Weightage / Month", f"Labour Cost ({currency})"]
    if grid_df.empty: return pd.DataFrame(columns=out_cols)
    required = ["Personnel", "Category", "Type of Unit Rate", "Total Hours", "Weightage (FTE)"]
    if not all(c in grid_df.columns for c in required):
        return pd.DataFrame(columns=out_cols)
    cache = {}
    def rate_for(row):
        key = (row["Personnel"], row["Category"], row["Type of Unit Rate"],
               type_of_schedule, type_of_package)
        if key in cache: return cache[key]
        val = get_rate(mec_df, *key)
        cache[key] = val
        return val
    df = grid_df.copy()
    if "Weightage / Month" not in df.columns:
        df["Weightage / Month"] = 0.0
    df["Total Hours"]     = pd.to_numeric(df["Total Hours"],     errors="coerce").fillna(0).astype(float)
    df["Weightage (FTE)"] = pd.to_numeric(df["Weightage (FTE)"], errors="coerce").fillna(0).astype(float)
    df[f"Unit Rate ({currency})"]    = df.apply(rate_for, axis=1).astype(float)
    df[f"Labour Cost ({currency})"]  = df["Total Hours"] * df[f"Unit Rate ({currency})"]
    return df[out_cols]

def parse_lumpsum_details(raw):
    """Parse a JSON list of {detail, amount} dicts from a stored cell value."""
    if not raw:
        return [{"detail": "", "amount": 0.0}]
    if isinstance(raw, list):
        return raw
    try:
        parsed = json.loads(str(raw))
        if isinstance(parsed, list) and all(isinstance(it, dict) for it in parsed):
            return parsed
    except Exception:
        pass
    # Legacy plain-string — migrate to list format
    return [{"detail": str(raw), "amount": 0.0}]

def calculate_third_party_costs(df, total_labour, currency):
    if df.empty: return pd.DataFrame(columns=["Category", "Description", "Details", "Basis", f"Cost ({currency})"])
    result = df.copy()
    result[f"Cost ({currency})"] = 0.0
    for idx, row in result.iterrows():
        basis = str(row["Basis"]).lower()
        if "percentage" in basis or "%" in basis:
            result.loc[idx, f"Cost ({currency})"] = total_labour * (float(row.get("Percentage", 0)) / 100.0)
        else:
            result.loc[idx, f"Cost ({currency})"] = float(row.get("Fixed Amount", 0))

    def fmt_details(row):
        raw = row.get("LumpSum Details", "")
        if not raw:
            return ""
        try:
            items = json.loads(str(raw))
            if isinstance(items, list):
                parts = [f"{it.get('detail','')}" for it in items if it.get("detail") or it.get("amount")]
                return "; ".join(parts) if parts else ""
        except Exception:
            pass
        return str(raw)

    result["Details"] = result.apply(fmt_details, axis=1)
    return result[["Category", "Description", "Details", "Basis", f"Cost ({currency})"]]

def apply_monthly_loading(labour_df, third_party_df, monthly_df, currency):
    if monthly_df.empty:
        return {"monthly_labour": {}, "monthly_third_party": {}, "total_by_month": {}, "months": []}
    months  = monthly_df["Month"].tolist()
    factors = monthly_df["Loading Factor (%)"].tolist()
    weights = monthly_df["Weightage Distribution"].tolist()
    total_labour = labour_df[f"Labour Cost ({currency})"].sum() if not labour_df.empty else 0
    total_third  = third_party_df[f"Cost ({currency})"].sum() if not third_party_df.empty else 0
    ml, mt, tb = {}, {}, {}
    for i, m in enumerate(months):
        factor = factors[i] / 100.0
        weight = weights[i] / 100.0
        ml[m] = total_labour * weight * factor
        mt[m] = total_third  * weight * factor
        tb[m] = ml[m] + mt[m]
    return {"monthly_labour": ml, "monthly_third_party": mt, "total_by_month": tb, "months": months}

def compute_totals(labour_df, third_party_df, monthly_df, currency):
    total_labour = labour_df[f"Labour Cost ({currency})"].sum() if not labour_df.empty else 0.0
    total_hours  = labour_df["Total Hours"].sum()               if not labour_df.empty else 0.0
    total_third  = third_party_df[f"Cost ({currency})"].sum()   if not third_party_df.empty else 0.0
    total_exact  = total_labour + total_third
    if not labour_df.empty:
        disc_totals = labour_df.groupby("Discipline", as_index=False).agg(
            Manhour=("Total Hours", "sum"),
            **{f"Labour Cost ({currency})": (f"Labour Cost ({currency})", "sum")}
        )
    else:
        disc_totals = pd.DataFrame(columns=["Discipline", "Manhour", f"Labour Cost ({currency})"])
    monthly = apply_monthly_loading(labour_df, third_party_df, monthly_df, currency)
    return {"total_hours": total_hours, "total_labour_exact": total_labour,
            "total_third_party": total_third, "total_exact": total_exact,
            "discipline_totals": disc_totals, "monthly_breakdown": monthly}

# ─────────────────────────────────────────────────
# Save / compare
def save_current_project():
    sched = st.session_state["type_of_schedule"]
    pkg   = st.session_state["type_of_package"]
    curr  = currency_for(sched)
    labour = calculate_labour_costs(st.session_state[GRID_KEY], curr, sched, pkg)
    total_lab = labour[f"Labour Cost ({curr})"].sum() if not labour.empty else 0.0
    combined = pd.concat([st.session_state[THIRD_PARTY_KEY], st.session_state.get(NON_LABOUR_KEY, pd.DataFrame())], ignore_index=True)
    third  = calculate_third_party_costs(combined, total_lab, curr)
    totals = compute_totals(labour, third, st.session_state[MONTHLY_LOADING_KEY], curr)
    data = {
        "id": datetime.now().strftime("%Y%m%d_%H%M%S"),
        "name": st.session_state["project_title"] or f"Project_{datetime.now():%Y%m%d_%H%M}",
        "timestamp": datetime.now().isoformat(),
        "project_title": st.session_state["project_title"],
        "type_of_schedule": sched, "type_of_package": pkg, "currency": curr,
        "total_hours": totals["total_hours"],
        "total_labour_exact": totals["total_labour_exact"],
        "total_third_party": totals["total_third_party"],
        "total_exact": totals["total_exact"],
        "discipline_totals": totals["discipline_totals"].to_dict("records") if not totals["discipline_totals"].empty else [],
        "labour_line_items": labour.to_dict("records"),
        "third_party_items": third.to_dict("records"),
        "personnel_count": len(st.session_state[GRID_KEY]),
        "disciplines_used": st.session_state[GRID_KEY]["Discipline"].nunique()
    }
    st.session_state[SAVED_PROJECTS_KEY].append(data)
    return data

def compare_projects(p1, p2):
    return {
        "hours_diff": p2["total_hours"] - p1["total_hours"],
        "hours_pct": ((p2["total_hours"] - p1["total_hours"]) / p1["total_hours"] * 100) if p1["total_hours"] else 0,
        "exact_cost_diff": p2["total_exact"] - p1["total_exact"],
        "exact_cost_pct": ((p2["total_exact"] - p1["total_exact"]) / p1["total_exact"] * 100) if p1["total_exact"] else 0,
    }

# ─────────────────────────────────────────────────
# Excel export
def to_excel_bytes(main_meta, totals, labour_df, third_party_df, monthly_df, currency):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("B2:I3")
    ws["B2"] = "MAJOR ENGINEERING CONTRACT (MEC) TOOL FOR CE UPSTREAM"
    ws["B2"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws["B2"].fill  = PatternFill("solid", fgColor="00A19C")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].border = border_all
    row = 5
    for k, v in main_meta.iloc[0].items():
        ws[f"B{row}"] = k
        ws[f"C{row}"] = v
        ws[f"B{row}"].border = border_all
        ws[f"C{row}"].border = border_all
        ws[f"B{row}"].font = Font(bold=True)
        row += 1
    row += 2
    for lbl, col in [("Description","B"),("Manhour","C"),(f"Total Price ({currency})","D")]:
        ws[f"{col}{row}"] = lbl
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill("solid", fgColor="00A19C")
        ws[f"{col}{row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col}{row}"].border = border_all
    row += 1
    if not labour_df.empty:
        disc_sum = labour_df.groupby("Discipline").agg({"Total Hours":"sum",f"Labour Cost ({currency})":"sum"}).reset_index()
        for _, drow in disc_sum.iterrows():
            ws[f"B{row}"] = drow["Discipline"]
            ws[f"C{row}"] = drow["Total Hours"]
            ws[f"D{row}"] = drow[f"Labour Cost ({currency})"]
            for col in ["B", "C", "D"]:
                ws[f"{col}{row}"].border = border_all
            row += 1
    ws[f"B{row}"] = "B Third Party Services Cost(*)"
    ws[f"D{row}"] = totals["total_third_party"]
    for col in ["B", "C", "D"]:
        ws[f"{col}{row}"].border = border_all
    row += 1
    non_labour_df = st.session_state.get(NON_LABOUR_KEY, pd.DataFrame())
    if not non_labour_df.empty:
        non_cost = calculate_third_party_costs(non_labour_df, totals["total_labour_exact"], currency)[f"Cost ({currency})"].sum()
        ws[f"B{row}"] = "C Non-Labour Cost"
        ws[f"D{row}"] = non_cost
        for col in ["B", "C", "D"]:
            ws[f"{col}{row}"].border = border_all
        row += 1
    row += 1
    ws[f"B{row}"] = "Total Raw Bid Price (Base Scope)"
    ws[f"C{row}"] = totals["total_hours"]
    ws[f"D{row}"] = totals["total_exact"]
    for col in ["B","C","D"]:
        ws[f"{col}{row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col}{row}"].fill = PatternFill("solid", fgColor="00A19C")
        ws[f"{col}{row}"].border = border_all
    row += 2
    total_weightage = labour_df["Weightage (FTE)"].sum() if not labour_df.empty else 0
    num_months_excel = len(monthly_df) if not monthly_df.empty else N_MONTHS
    avg_weight = total_weightage / num_months_excel if num_months_excel else 0
    ws[f"B{row}"] = "Average Weightage per Month"
    ws[f"C{row}"] = round(avg_weight, 2)
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].border = border_all
    ws[f"C{row}"].border = border_all
    row += 1
    rate_per_mh = totals['total_exact'] / totals['total_hours'] if totals['total_hours'] else 0
    ws[f"B{row}"] = f"{currency}/Manhour"
    ws[f"C{row}"] = round(rate_per_mh, 2)
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"B{row}"].border = border_all
    ws[f"C{row}"].border = border_all
    if not labour_df.empty:
        ws2 = wb.create_sheet("Labour Details")
        for c, col in enumerate(labour_df.columns, 1):
            ws2.cell(1, c, col).font = Font(bold=True)
            ws2.cell(1, c, col).fill = PatternFill("solid", fgColor="E6F7F7")
        for r, row_data in labour_df.iterrows():
            for c, val in enumerate(row_data, 1):
                ws2.cell(r+2, c, val)
    # Set explicit column widths so cells aren't smashed
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 4
    if "Labour Details" in wb.sheetnames:
        ws2 = wb["Labour Details"]
        for i, col in enumerate(labour_df.columns, 1):
            max_len = max(len(str(col)), labour_df[col].astype(str).map(len).max() if not labour_df.empty else 0)
            ws2.column_dimensions[chr(64 + i)].width = min(max_len + 4, 40)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ─────────────────────────────────────────────────
# Step indicator helper
PAGE_STEPS = ["MAIN", "TABLE", "THIRD_PARTY", "NON_LABOUR", "LOADING", "TOTALS", "SUMMARY"]

def render_step_bar():
    current = st.session_state["page"]
    labels  = {"MAIN":"Setup","TABLE":"Personnel","THIRD_PARTY":"Third Party",
               "NON_LABOUR":"Non-Labour","LOADING":"Loading","TOTALS":"Totals","SUMMARY":"Summary"}
    html = '<div class="step-bar">'
    for i, step in enumerate(PAGE_STEPS):
        if i > 0:
            html += '<div class="step-sep"></div>'
        done   = PAGE_STEPS.index(current) > i if current in PAGE_STEPS else False
        active = step == current
        cls    = "active" if active else ("done" if done else "")
        icon   = "✓" if done else str(i+1)
        html += f'''
        <div class="step-item {'active' if active else ''}">
            <div class="step-dot {cls}">{icon}</div>
            <span>{labels[step]}</span>
        </div>'''
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# Navigation bar (no emojis)
def render_navigation():
    st.markdown('<div class="nav-marker"></div>', unsafe_allow_html=True)
    page_order = list(PAGES.items())
    n = len(page_order)
    cols = st.columns(n, gap="small")
    for i, (key, label) in enumerate(page_order):
        with cols[i]:
            is_active = st.session_state["page"] == key
            btn_type = "primary" if is_active else "secondary"
            if st.button(label, key=f"nav_{key}", use_container_width=True, type=btn_type):
                st.session_state["page"] = key
                st.rerun()
    st.markdown("<hr style='border:none;border-top:1px solid var(--border);margin:0.25rem 0 1.5rem 0'>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# Section header helper (no emojis)
def section_header(title, icon=""):
    st.markdown(f"""
    <div class="section-header fadeup">
        <div class="dot"></div>
        <h3>{icon + '  ' if icon else ''}{title}</h3>
    </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# Page: MAIN (no emojis)
def render_main():
    # Banner
    data_status = ""
    if st.session_state["mec_data_loaded"]:
        data_status += f'<span class="badge pulse">● MEC Loaded — {len(mec_df)} rows</span>'
    if st.session_state["mdr_data_loaded"]:
        data_status += f'<span class="badge" style="margin-left:0.5rem">● MDR Loaded — {len(mdr_df)} rows</span>'

    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>MEC Tool</h1>
        <p class="subtitle">Major Engineering Contract Cost Estimation — PETRONAS Upstream CE</p>
        {data_status}
    </div>
    """, unsafe_allow_html=True)

    col_reset = st.columns([6, 1])
    with col_reset[1]:
        if st.button("New Project", use_container_width=True):
            reset_all(); st.rerun()

    render_step_bar()

    section_header("Project Information", "📋")
    mp1, mp2 = st.columns(2, gap="large")

    with mp1:
        st.session_state["project_title"]  = st.text_input("Project Title",    st.session_state["project_title"])
        st.session_state["cost_engineer"]  = st.text_input("Cost Engineer",    st.session_state["cost_engineer"])
        st.session_state["tp_specialist"]  = st.text_input("TP / Specialist",  st.session_state["tp_specialist"])

    with mp2:
        st.session_state["project_date"] = st.date_input("Project Date", st.session_state["project_date"])
        if schedule_opts:
            cur = st.session_state["type_of_schedule"]
            idx = schedule_opts.index(cur) if cur in schedule_opts else 0
            st.session_state["type_of_schedule"] = st.selectbox("Schedule", schedule_opts, index=idx)
        else:
            st.session_state["type_of_schedule"] = st.selectbox("Schedule", ["A","B","C","D"])
        st.session_state["type_of_package"] = st.selectbox(
            "Package", ["U1","U2"],
            index=0 if st.session_state["type_of_package"] == "U1" else 1
        )

    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)

    sched = st.session_state["type_of_schedule"]
    curr  = currency_for(sched)
    m1, m2, m3 = st.columns(3, gap="medium")
    with m1:
        st.markdown(f"""
        <div class="metric-card fadeup">
            <div class="label">Currency</div>
            <div class="value">{curr}</div>
            <div class="sub">Schedule {sched}</div>
        </div>""", unsafe_allow_html=True)
    with m2:
        st.markdown(f"""
        <div class="metric-card fadeup">
            <div class="label">Hours / Month</div>
            <div class="value">{int(HOURS_PER_MONTH)}</div>
            <div class="sub">Standard working hours</div>
        </div>""", unsafe_allow_html=True)
    with m3:
        st.markdown(f"""
        <div class="metric-card fadeup">
            <div class="label">Project Duration</div>
            <div class="value">{N_MONTHS} mo</div>
            <div class="sub">Fixed project duration</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    b1, b2 = st.columns([1, 1], gap="small")
    with b1:
        if st.button("Reset Grid", use_container_width=True):
            reset_grid(); st.rerun()
    with b2:
        if st.button("Next →", use_container_width=True, type="primary"):
            st.session_state["page"] = "TABLE"; st.rerun()

# ─────────────────────────────────────────────────
# Page: TABLE (no emojis)
def render_table():
    sched = st.session_state["type_of_schedule"]
    pkg   = st.session_state["type_of_package"]
    curr  = currency_for(sched)
    cats  = build_category_options(sched, mec_df) or (B_D_CATEGORY_FALLBACK if is_usd(sched) else AC_CATEGORIES)

    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>Personnel Table</h1>
        <p class="subtitle">Configure discipline resources — Schedule {sched} · Package {pkg} · {curr}</p>
    </div>
    """, unsafe_allow_html=True)

    render_step_bar()

    df = st.session_state[GRID_KEY].copy()
    if df.empty:
        initialize_default_grids()
        df = st.session_state[GRID_KEY].copy()
    if not df.empty and "Weightage / Month" not in df.columns:
        df["Weightage / Month"] = 0.0
    df["Category"] = df["Category"].where(df["Category"].isin(cats), cats[0] if cats else "")
    df["Swatch"]   = df["Discipline"].map(DISCIPLINE_SWATCH).fillna("⬜")
    st.session_state[GRID_KEY] = df

    # Bulk actions
    section_header("Bulk Actions", "⚡")
    _bleft, _bgap, _bright = st.columns([3, 0.3, 3])
    with _bleft:
        bc1, bc2 = st.columns([3, 1], gap="small", vertical_alignment="bottom")
        with bc1:
            st.selectbox("Category for ALL rows", cats, key="bulk_cat")
        with bc2:
            if st.button("Apply", use_container_width=True, key="apply_cat"):
                df["Category"] = st.session_state["bulk_cat"]
                st.session_state[GRID_KEY] = df; st.rerun()
    with _bright:
        bc3, bc4 = st.columns([3, 1], gap="small", vertical_alignment="bottom")
        with bc3:
            st.selectbox("Rate Type for ALL rows", RATE_TYPES or UNIT_TYPES, key="bulk_type")
        with bc4:
            if st.button("Apply", use_container_width=True, key="apply_type"):
                df["Type of Unit Rate"] = st.session_state["bulk_type"]
                st.session_state[GRID_KEY] = df; st.rerun()

    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
    section_header("Personnel Grid", "📋")

    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(resizable=True, sortable=True, filter=True)
    gb.configure_column("Swatch",           pinned="left",  width=60,  editable=False)
    gb.configure_column("Discipline",       pinned="left",  width=190, editable=True,
                        cellEditor="agSelectCellEditor",
                        cellEditorParams={"values": list(DISCIPLINE_ROW_COUNTS.keys())})
    gb.configure_column("Personnel",        width=260, editable=True,
                        cellEditor="agSelectCellEditor",
                        cellEditorParams={"values": PERSONNEL_LIST or ["Project Manager"]})
    gb.configure_column("Category",         width=240, editable=True,
                        cellEditor="agSelectCellEditor", cellEditorParams={"values": cats})
    gb.configure_column("Type of Unit Rate",width=200, editable=True,
                        cellEditor="agSelectCellEditor",
                        cellEditorParams={"values": RATE_TYPES or UNIT_TYPES})
    gb.configure_column("Total Hours",      type=["numericColumn"], width=150, editable=False)
    gb.configure_column("Weightage (FTE)",  type=["numericColumn"], width=150, editable=False)
    gb.configure_column("Weightage / Month",type=["numericColumn"], width=150, editable=False)

    disc_bg  = {k: f"#{v}" for k, v in DISCIPLINE_COLORS.items()}
    row_style = JsCode(f"""
        function(params) {{
            const map = {json.dumps(disc_bg)};
            return params.data ? {{backgroundColor: map[params.data.Discipline] || null}} : null;
        }}
    """)
    gb.configure_grid_options(
        getRowStyle=row_style,
        onGridSizeChanged=JsCode("function(params){params.api.sizeColumnsToFit();}"),
        onFirstDataRendered=JsCode("function(params){params.api.sizeColumnsToFit();}"),
    )

    try:
        ag_theme = "alpine-dark" if st.session_state.get("dark_mode", False) else "alpine"
        resp = AgGrid(df, gridOptions=gb.build(), height=480, update_on="value_changed",
                      allow_unsafe_jscode=True, theme=ag_theme, fit_columns_on_grid_load=True)
        st.session_state[GRID_KEY] = pd.DataFrame(resp["data"])
    except Exception as e:
        st.warning(f"Grid error: {e}")

    b1, b2, b3, b4 = st.columns(4, gap="small")
    with b1:
        if st.button("← Back", use_container_width=True):
            st.session_state["page"] = "MAIN"; st.rerun()
    with b2:
        if st.button("Add Row", use_container_width=True):
            df = st.session_state[GRID_KEY].copy()
            df.loc[len(df)] = {
                "Swatch": "⬜", "Discipline": list(DISCIPLINE_ROW_COUNTS.keys())[0],
                "Personnel": PERSONNEL_LIST[0] if PERSONNEL_LIST else "",
                "Category": cats[0] if cats else "", "Type of Unit Rate": "Normalise",
                "Total Hours": 0.0, "Weightage (FTE)": 0.0, "Weightage / Month": 0.0
            }
            st.session_state[GRID_KEY] = df; st.rerun()
    with b3:
        if st.button("Reset", use_container_width=True):
            reset_grid(); st.rerun()
    with b4:
        if st.button("Next →", use_container_width=True, type="primary"):
            st.session_state["page"] = "THIRD_PARTY"; st.rerun()

    # Live total
    labour = calculate_labour_costs(st.session_state[GRID_KEY], curr, sched, pkg)
    total  = labour[f"Labour Cost ({curr})"].sum()
    total_h = labour["Total Hours"].sum() if not labour.empty else 0
    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
    lc1, lc2 = st.columns(2)
    with lc1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Total Labour Cost</div>
            <div class="value">{curr} {total:,.2f}</div>
        </div>""", unsafe_allow_html=True)
    with lc2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Total Manhours</div>
            <div class="value">{total_h:,.0f} hrs</div>
        </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# Page: TOTALS (no emojis)
def render_totals():
    sched = st.session_state["type_of_schedule"]
    pkg   = st.session_state["type_of_package"]
    curr  = currency_for(sched)

    labour   = calculate_labour_costs(st.session_state[GRID_KEY], curr, sched, pkg)
    total_lab = labour[f"Labour Cost ({curr})"].sum() if not labour.empty else 0.0
    combined  = pd.concat([st.session_state[THIRD_PARTY_KEY], st.session_state.get(NON_LABOUR_KEY, pd.DataFrame())], ignore_index=True)
    third    = calculate_third_party_costs(combined, total_lab, curr)
    totals   = compute_totals(labour, third, st.session_state[MONTHLY_LOADING_KEY], curr)

    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>Project Totals</h1>
        <p class="subtitle">Cost summary — Schedule {sched} · Package {pkg} · {curr}</p>
    </div>
    """, unsafe_allow_html=True)

    render_step_bar()

    m1, m2, m3, m4 = st.columns(4, gap="medium")
    with m1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Total Manhours</div>
            <div class="value">{totals['total_hours']:,.0f}</div>
            <div class="sub">hrs</div>
        </div>""", unsafe_allow_html=True)
    with m2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Labour Cost</div>
            <div class="value">{curr} {totals['total_labour_exact']:,.0f}</div>
        </div>""", unsafe_allow_html=True)
    with m3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Third Party + Non-Labour</div>
            <div class="value">{curr} {totals['total_third_party']:,.0f}</div>
        </div>""", unsafe_allow_html=True)
    with m4:
        st.markdown(f"""
        <div class="metric-card" style="border-top-color:var(--gold)">
            <div class="label">Total Bid Price</div>
            <div class="value" style="color:var(--gold)">{curr} {totals['total_exact']:,.0f}</div>
        </div>""", unsafe_allow_html=True)

    if not totals["discipline_totals"].empty:
        st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
        section_header("Labour Cost by Discipline", "🏗️")
        disc = totals["discipline_totals"].copy()
        disc[f"Labour Cost ({curr})"] = disc[f"Labour Cost ({curr})"].apply(lambda x: f"{curr} {x:,.2f}")
        disc["Manhour"] = disc["Manhour"].apply(lambda x: f"{x:,.0f}")
        st.dataframe(disc, use_container_width=True, hide_index=True)

    b1, b2, b3, b4 = st.columns(4, gap="small")
    with b1:
        if st.button("Back", use_container_width=True):
            st.session_state["page"] = "LOADING"; st.rerun()
    with b2:
        if st.button("Save Project", use_container_width=True):
            save_current_project(); st.success("Project saved!"); st.rerun()
    with b3:
        if st.button("Dashboard", use_container_width=True, type="primary"):
            st.session_state["page"] = "SUMMARY"; st.rerun()
    with b4:
        if st.button("Compare", use_container_width=True):
            st.session_state["page"] = "COMPARE"; st.rerun()

# ─────────────────────────────────────────────────
# Page: SUMMARY (no emojis)
def render_summary():
    sched = st.session_state["type_of_schedule"]
    pkg   = st.session_state["type_of_package"]
    curr  = currency_for(sched)

    meta = pd.DataFrame([{
        "Project Title": st.session_state["project_title"],
        "Date": str(st.session_state["project_date"]),
        "Cost Engineer": st.session_state["cost_engineer"],
        "TP/Specialist": st.session_state["tp_specialist"],
        "Schedule": sched, "Package": pkg, "Currency": curr,
        "Hours/Month": HOURS_PER_MONTH
    }])

    labour    = calculate_labour_costs(st.session_state[GRID_KEY], curr, sched, pkg)
    total_lab = labour[f"Labour Cost ({curr})"].sum() if not labour.empty else 0.0
    combined  = pd.concat([st.session_state[THIRD_PARTY_KEY], st.session_state.get(NON_LABOUR_KEY, pd.DataFrame())], ignore_index=True)
    third     = calculate_third_party_costs(combined, total_lab, curr)
    totals    = compute_totals(labour, third, st.session_state[MONTHLY_LOADING_KEY], curr)

    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>Project Dashboard</h1>
        <p class="subtitle">{st.session_state['project_title'] or 'Untitled Project'} — {sched} · {pkg} · {curr}</p>
    </div>
    """, unsafe_allow_html=True)

    render_step_bar()

    # KPI row
    total_weightage = labour["Weightage (FTE)"].sum() if not labour.empty else 0
    monthly_df = st.session_state.get(MONTHLY_LOADING_KEY, pd.DataFrame())
    num_dash_months = len(monthly_df) if not monthly_df.empty else N_MONTHS
    avg_weight      = total_weightage / num_dash_months if num_dash_months else 0
    rate_per_mh     = totals["total_exact"] / totals["total_hours"] if totals["total_hours"] else 0

    k1, k2, k3, k4 = st.columns(4, gap="medium")
    with k1:
        st.markdown(f"""<div class="metric-card"><div class="label">Total Manhours</div>
        <div class="value">{totals['total_hours']:,.0f}</div></div>""", unsafe_allow_html=True)
    with k2:
        st.markdown(f"""<div class="metric-card"><div class="label">Avg Weightage/Month</div>
        <div class="value">{avg_weight:.2f} FTE</div></div>""", unsafe_allow_html=True)
    with k3:
        st.markdown(f"""<div class="metric-card"><div class="label">{curr}/Manhour</div>
        <div class="value">{rate_per_mh:,.2f}</div></div>""", unsafe_allow_html=True)
    with k4:
        st.markdown(f"""<div class="metric-card" style="border-top-color:var(--gold)"><div class="label">Total Bid Price</div>
        <div class="value" style="color:var(--gold)">{curr} {totals['total_exact']:,.0f}</div></div>""", unsafe_allow_html=True)

    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)

    # Charts
    if not labour.empty:
        ch1, ch2 = st.columns(2, gap="large")
        disc_sum = labour.groupby("Discipline").agg(
            {"Total Hours": "sum", f"Labour Cost ({curr})": "sum"}).reset_index()

        with ch1:
            section_header("Labour Cost by Discipline", "💹")
            fig = px.bar(
                disc_sum, x=f"Labour Cost ({curr})", y="Discipline",
                orientation="h", color=f"Labour Cost ({curr})",
                color_continuous_scale=[[0, PETRONAS["teal_deep"]], [0.5, PETRONAS["teal"]], [1, PETRONAS["gold"]]],
                template="plotly_dark" if st.session_state["dark_mode"] else "plotly_white"
            )
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=0,r=0,t=10,b=0), coloraxis_showscale=False,
                font=dict(family="DM Sans"), yaxis=dict(categoryorder="total ascending")
            )
            st.plotly_chart(fig, use_container_width=True)

        with ch2:
            section_header("Manhour Distribution", "⏱️")
            fig2 = px.pie(
                disc_sum, values="Total Hours", names="Discipline",
                color_discrete_sequence=px.colors.sequential.Teal,
                template="plotly_dark" if st.session_state["dark_mode"] else "plotly_white",
                hole=0.45
            )
            fig2.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=0,r=0,t=10,b=0),
                legend=dict(font=dict(size=10, family="DM Sans")),
                font=dict(family="DM Sans")
            )
            st.plotly_chart(fig2, use_container_width=True)

    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
    section_header("Project Information", "📋")
    st.dataframe(meta, use_container_width=True, hide_index=True)

    section_header("Cost Summary", "💰")
    summary_rows = []
    if not labour.empty:
        disc_sum2 = labour.groupby("Discipline").agg({"Total Hours":"sum", f"Labour Cost ({curr})":"sum"}).reset_index()
        for _, row in disc_sum2.iterrows():
            summary_rows.append({
                "Description": row["Discipline"],
                "Manhour": f"{row['Total Hours']:,.0f}",
                f"Total Price ({curr})": f"{row[f'Labour Cost ({curr})']:,.2f}"
            })
    tp_cost = third[f"Cost ({curr})"].sum() if not third.empty else 0.0
    summary_rows.append({"Description":"B Third Party Services Cost(*)","Manhour":"",
                         f"Total Price ({curr})":f"{tp_cost:,.2f}"})
    non_labour_df = st.session_state.get(NON_LABOUR_KEY, pd.DataFrame())
    if not non_labour_df.empty:
        non_cost = calculate_third_party_costs(non_labour_df, total_lab, curr)[f"Cost ({curr})"].sum()
        summary_rows.append({"Description":"C Non-Labour Cost","Manhour":"",
                             f"Total Price ({curr})":f"{non_cost:,.2f}"})
    summary_rows.append({
        "Description": "Total Raw Bid Price (Base Scope)",
        "Manhour": f"{totals['total_hours']:,.0f}",
        f"Total Price ({curr})": f"{totals['total_exact']:,.2f}"
    })
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
    b1, b2, b3, b4 = st.columns(4, gap="small")
    with b1:
        if st.button("Back", use_container_width=True):
            st.session_state["page"] = "TOTALS"; st.rerun()
    with b2:
        if st.button("Save Project", use_container_width=True):
            save_current_project(); st.success("Saved!"); st.rerun()
    with b3:
        if st.button("Compare", use_container_width=True):
            st.session_state["page"] = "COMPARE"; st.rerun()
    with b4:
        excel = to_excel_bytes(meta, totals, labour, third, st.session_state[MONTHLY_LOADING_KEY], curr)
        st.download_button(
            "Excel Report", data=excel,
            file_name=f"MEC_{st.session_state['project_title'] or 'Output'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary"
        )

# ─────────────────────────────────────────────────
# Page: THIRD PARTY (no emojis)
def render_third_party():
    currency = currency_for(st.session_state["type_of_schedule"])
    labour_df = calculate_labour_costs(st.session_state[GRID_KEY], currency,
                                       st.session_state["type_of_schedule"],
                                       st.session_state["type_of_package"])
    total_labour = labour_df[f"Labour Cost ({currency})"].sum() if not labour_df.empty else 0

    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>Third Party Services</h1>
        <p class="subtitle">Define third-party costs — base labour: {currency} {total_labour:,.2f}</p>
    </div>
    """, unsafe_allow_html=True)

    render_step_bar()

    df = st.session_state[THIRD_PARTY_KEY].copy()
    if df.empty:
        df = pd.DataFrame([{"Category":"Third Party Services","Description":"",
                            "Basis":"Percentage of Labour Cost","Percentage":0.0,
                            "Fixed Amount":0.0,"Remarks":"","LumpSum Details":""}])
        st.session_state[THIRD_PARTY_KEY] = df
    # Ensure LumpSum Details column exists
    if "LumpSum Details" not in df.columns:
        df["LumpSum Details"] = ""

    df["Basis"] = df["Basis"].replace({"% of Labour Cost":"Percentage of Labour Cost",
                                       "Fixed Amount":"LumpSum / Fixed Amount"})

    # Handle pending remove-detail action (idx, d_i)
    _rem_det = st.session_state.pop("tp_rem_det", None)
    if _rem_det is not None:
        _ridx, _rdi = _rem_det
        if _ridx in df.index:
            _items = parse_lumpsum_details(df.loc[_ridx, "LumpSum Details"])
            if len(_items) > 1:
                _items.pop(_rdi)
                df.loc[_ridx, "LumpSum Details"] = json.dumps(_items)
                df.loc[_ridx, "Fixed Amount"] = sum(it["amount"] for it in _items)
            st.session_state[THIRD_PARTY_KEY] = df; st.rerun()

    remove_idx = st.session_state.pop("tp_remove_idx", None)
    if remove_idx is not None and remove_idx in df.index:
        df = df.drop(index=remove_idx).reset_index(drop=True)
        st.session_state[THIRD_PARTY_KEY] = df; st.rerun()

    basis_options = ["Percentage of Labour Cost", "LumpSum / Fixed Amount"]
    for idx, row in df.iterrows():
        current_basis = row["Basis"] if row["Basis"] in basis_options else basis_options[0]
        col1, col2, col3, col4 = st.columns([3, 2, 2, 1], gap="small")
        with col1:
            df.loc[idx, "Description"] = st.text_input(f"Description", value=row.get("Description", ""),
                key=f"tp_desc_{idx}", placeholder="Enter description…")
        with col2:
            df.loc[idx, "Basis"] = st.selectbox("Basis", basis_options,
                index=basis_options.index(current_basis), key=f"tp_basis_{idx}")
        with col3:
            if df.loc[idx, "Basis"] == "Percentage of Labour Cost":
                df.loc[idx, "Percentage"] = st.number_input("Percentage (%)",
                    min_value=0.0, max_value=100.0, value=float(row.get("Percentage", 0.0)),
                    step=0.0001, format="%.4f", key=f"tp_pct_{idx}")
                df.loc[idx, "Fixed Amount"] = 0.0
                df.loc[idx, "LumpSum Details"] = ""
            # For LumpSum, col3 is left blank intentionally
        with col4:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️", key=f"tp_del_{idx}"):
                st.session_state["tp_remove_idx"] = idx; st.rerun()

        if df.loc[idx, "Basis"] == "LumpSum / Fixed Amount":
            df.loc[idx, "Percentage"] = 0.0
            detail_items = parse_lumpsum_details(row.get("LumpSum Details", ""))

            # Render each detail row
            for d_i, det in enumerate(detail_items):
                _, d_col1, d_col2, d_col3 = st.columns([3, 2, 2, 1], gap="small")
                with d_col1:
                    det["detail"] = st.text_input("Detail", value=det.get("detail", ""),
                        key=f"tp_ls_det_{idx}_{d_i}", placeholder="Enter detail…")
                with d_col2:
                    det["amount"] = st.number_input("Amount",
                        min_value=0.0, value=float(det.get("amount", 0.0)),
                        step=100.0, format="%.2f", key=f"tp_ls_amt_{idx}_{d_i}")
                with d_col3:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if len(detail_items) > 1 and st.button("✕", key=f"tp_ls_rem_{idx}_{d_i}"):
                        st.session_state["tp_rem_det"] = (idx, d_i)
                        df.loc[idx, "LumpSum Details"] = json.dumps(detail_items)
                        df.loc[idx, "Fixed Amount"] = sum(it["amount"] for it in detail_items)
                        st.session_state[THIRD_PARTY_KEY] = df; st.rerun()

            # Total Amount row
            total_det = sum(it["amount"] for it in detail_items)
            _, ta_col1, ta_col2, _ = st.columns([3, 2, 2, 1], gap="small")
            with ta_col1:
                st.markdown(
                    "<div style='text-align:right;font-weight:700;color:var(--text);"
                    "border-top:2px solid var(--border);padding:6px 4px 2px'>Total Amount</div>",
                    unsafe_allow_html=True)
            with ta_col2:
                st.markdown(
                    f"<div style='font-weight:700;color:var(--text);"
                    f"border-top:2px solid var(--border);padding:6px 0 2px'>{total_det:,.2f}</div>",
                    unsafe_allow_html=True)

            # Add Detail button
            _, ad_col, _, _ = st.columns([3, 2, 2, 1], gap="small")
            with ad_col:
                if st.button("+ Add Detail", key=f"tp_add_det_{idx}", use_container_width=True):
                    detail_items.append({"detail": "", "amount": 0.0})
                    df.loc[idx, "LumpSum Details"] = json.dumps(detail_items)
                    df.loc[idx, "Fixed Amount"] = total_det
                    st.session_state[THIRD_PARTY_KEY] = df; st.rerun()

            # Sync back to df
            df.loc[idx, "LumpSum Details"] = json.dumps(detail_items)
            df.loc[idx, "Fixed Amount"] = total_det
        
        st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)

    st.session_state[THIRD_PARTY_KEY] = df

    col1, col2 = st.columns(2, gap="small")
    col1, col2, col3 = st.columns(3, gap="small")
    with col1:
        if st.button("← Back", use_container_width=True):
            st.session_state["page"] = "TABLE"; st.rerun()
    with col2:
        if st.button("Add Item", use_container_width=True):
            new = pd.DataFrame([{"Category":"Third Party Services","Description":"",
                                 "Basis":"Percentage of Labour Cost","Percentage":0.0,
                                 "Fixed Amount":0.0,"Remarks":"","LumpSum Details":""}])
            st.session_state[THIRD_PARTY_KEY] = pd.concat([df, new], ignore_index=True); st.rerun()
    with col3:
        if st.button("Next →", use_container_width=True, type="primary"):
            st.session_state["page"] = "NON_LABOUR"; st.rerun()

    if len(df) > 0:
        costs = calculate_third_party_costs(df, total_labour, currency)
        total = costs[f"Cost ({currency})"].sum()
        st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
        section_header("Calculated Costs", "🧮")
        st.dataframe(costs, use_container_width=True, hide_index=True)
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Total Third Party Cost</div>
            <div class="value">{currency} {total:,.2f}</div>
        </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# Page: NON-LABOUR (no emojis)
def render_non_labour():
    currency = currency_for(st.session_state["type_of_schedule"])
    labour_df = calculate_labour_costs(st.session_state[GRID_KEY], currency,
                                       st.session_state["type_of_schedule"],
                                       st.session_state["type_of_package"])
    total_labour = labour_df[f"Labour Cost ({currency})"].sum() if not labour_df.empty else 0

    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>Non-Labour Costs</h1>
        <p class="subtitle">Define non-labour costs — base labour: {currency} {total_labour:,.2f}</p>
    </div>
    """, unsafe_allow_html=True)

    render_step_bar()

    df = st.session_state[NON_LABOUR_KEY].copy()
    if df.empty:
        df = pd.DataFrame([{"Category":"Non-Labour Cost","Description":"",
                            "Basis":"Percentage of Labour Cost","Percentage":0.0,
                            "Fixed Amount":0.0,"Remarks":"","LumpSum Details":""}])
        st.session_state[NON_LABOUR_KEY] = df
    # Ensure LumpSum Details column exists
    if "LumpSum Details" not in df.columns:
        df["LumpSum Details"] = ""

    df["Basis"] = df["Basis"].replace({"% of Labour Cost":"Percentage of Labour Cost",
                                       "Fixed Amount":"LumpSum / Fixed Amount"})

    # Handle pending remove-detail action (idx, d_i)
    _rem_det = st.session_state.pop("nl_rem_det", None)
    if _rem_det is not None:
        _ridx, _rdi = _rem_det
        if _ridx in df.index:
            _items = parse_lumpsum_details(df.loc[_ridx, "LumpSum Details"])
            if len(_items) > 1:
                _items.pop(_rdi)
                df.loc[_ridx, "LumpSum Details"] = json.dumps(_items)
                df.loc[_ridx, "Fixed Amount"] = sum(it["amount"] for it in _items)
            st.session_state[NON_LABOUR_KEY] = df; st.rerun()

    remove_idx = st.session_state.pop("nl_remove_idx", None)
    if remove_idx is not None and remove_idx in df.index:
        df = df.drop(index=remove_idx).reset_index(drop=True)
        st.session_state[NON_LABOUR_KEY] = df; st.rerun()

    basis_options = ["Percentage of Labour Cost", "LumpSum / Fixed Amount"]
    for idx, row in df.iterrows():
        current_basis = row["Basis"] if row["Basis"] in basis_options else basis_options[0]
        col1, col2, col3, col4 = st.columns([3, 2, 2, 1], gap="small")
        with col1:
            df.loc[idx, "Description"] = st.text_input("Description", value=row.get("Description", ""),
                key=f"nl_desc_{idx}", placeholder="Enter description…")
        with col2:
            df.loc[idx, "Basis"] = st.selectbox("Basis", basis_options,
                index=basis_options.index(current_basis), key=f"nl_basis_{idx}")
        with col3:
            if df.loc[idx, "Basis"] == "Percentage of Labour Cost":
                df.loc[idx, "Percentage"] = st.number_input("Percentage (%)",
                    min_value=0.0, max_value=100.0, value=float(row.get("Percentage", 0.0)),
                    step=0.0001, format="%.4f", key=f"nl_pct_{idx}")
                df.loc[idx, "Fixed Amount"] = 0.0
                df.loc[idx, "LumpSum Details"] = ""
            # For LumpSum, col3 is left blank intentionally
        with col4:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️", key=f"nl_del_{idx}"):
                st.session_state["nl_remove_idx"] = idx; st.rerun()

        if df.loc[idx, "Basis"] == "LumpSum / Fixed Amount":
            df.loc[idx, "Percentage"] = 0.0
            detail_items = parse_lumpsum_details(row.get("LumpSum Details", ""))

            # Render each detail row
            for d_i, det in enumerate(detail_items):
                _, d_col1, d_col2, d_col3 = st.columns([3, 2, 2, 1], gap="small")
                with d_col1:
                    det["detail"] = st.text_input("Detail", value=det.get("detail", ""),
                        key=f"nl_ls_det_{idx}_{d_i}", placeholder="Enter detail…")
                with d_col2:
                    det["amount"] = st.number_input("Amount",
                        min_value=0.0, value=float(det.get("amount", 0.0)),
                        step=100.0, format="%.2f", key=f"nl_ls_amt_{idx}_{d_i}")
                with d_col3:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if len(detail_items) > 1 and st.button("✕", key=f"nl_ls_rem_{idx}_{d_i}"):
                        st.session_state["nl_rem_det"] = (idx, d_i)
                        df.loc[idx, "LumpSum Details"] = json.dumps(detail_items)
                        df.loc[idx, "Fixed Amount"] = sum(it["amount"] for it in detail_items)
                        st.session_state[NON_LABOUR_KEY] = df; st.rerun()

            # Total Amount row
            total_det = sum(it["amount"] for it in detail_items)
            _, ta_col1, ta_col2, _ = st.columns([3, 2, 2, 1], gap="small")
            with ta_col1:
                st.markdown(
                    "<div style='text-align:right;font-weight:700;color:var(--text);"
                    "border-top:2px solid var(--border);padding:6px 4px 2px'>Total Amount</div>",
                    unsafe_allow_html=True)
            with ta_col2:
                st.markdown(
                    f"<div style='font-weight:700;color:var(--text);"
                    f"border-top:2px solid var(--border);padding:6px 0 2px'>{total_det:,.2f}</div>",
                    unsafe_allow_html=True)

            # Add Detail button
            _, ad_col, _, _ = st.columns([3, 2, 2, 1], gap="small")
            with ad_col:
                if st.button("+ Add Detail", key=f"nl_add_det_{idx}", use_container_width=True):
                    detail_items.append({"detail": "", "amount": 0.0})
                    df.loc[idx, "LumpSum Details"] = json.dumps(detail_items)
                    df.loc[idx, "Fixed Amount"] = total_det
                    st.session_state[NON_LABOUR_KEY] = df; st.rerun()

            # Sync back to df
            df.loc[idx, "LumpSum Details"] = json.dumps(detail_items)
            df.loc[idx, "Fixed Amount"] = total_det

        st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)

    st.session_state[NON_LABOUR_KEY] = df

    col1, col2, col3 = st.columns(3, gap="small")
    with col1:
        if st.button("← Back", use_container_width=True):
            st.session_state["page"] = "THIRD_PARTY"; st.rerun()
    with col2:
        if st.button("Add Item", use_container_width=True):
            new = pd.DataFrame([{"Category":"Non-Labour Cost","Description":"",
                                 "Basis":"Percentage of Labour Cost","Percentage":0.0,
                                 "Fixed Amount":0.0,"Remarks":"","LumpSum Details":""}])
            st.session_state[NON_LABOUR_KEY] = pd.concat([df, new], ignore_index=True); st.rerun()
    with col3:
        if st.button("Next →", use_container_width=True, type="primary"):
            st.session_state["page"] = "LOADING"; st.rerun()

    if len(df) > 0:
        costs = calculate_third_party_costs(df, total_labour, currency)
        total = costs[f"Cost ({currency})"].sum()
        st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
        section_header("Calculated Costs", "🧮")
        st.dataframe(costs, use_container_width=True, hide_index=True)
        st.markdown(f"""
        <div class="metric-card">
            <div class="label">Total Non-Labour Cost</div>
            <div class="value">{currency} {total:,.2f}</div>
        </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# Page: LOADING (no emojis)
def render_loading():
    currency = currency_for(st.session_state["type_of_schedule"])
    labour_df = calculate_labour_costs(st.session_state[GRID_KEY], currency,
                                       st.session_state["type_of_schedule"],
                                       st.session_state["type_of_package"])
    total_labour = labour_df[f"Labour Cost ({currency})"].sum() if not labour_df.empty else 0

    combined = pd.concat([st.session_state[THIRD_PARTY_KEY], st.session_state.get(NON_LABOUR_KEY, pd.DataFrame())], ignore_index=True)
    third_party_df = calculate_third_party_costs(combined, total_labour, currency)

    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>Monthly Loading</h1>
        <p class="subtitle">Distribute costs across project duration — {currency} {total_labour:,.2f} base labour</p>
    </div>
    """, unsafe_allow_html=True)

    render_step_bar()

    df = st.session_state[MONTHLY_LOADING_KEY].copy()
    if df.empty:
        months = [f"Month {i+1:02d}" for i in range(N_MONTHS)]
        df = pd.DataFrame({"Month": months, "Loading Factor (%)": [100.0]*N_MONTHS,
                           "Weightage Distribution": [100.0/N_MONTHS]*N_MONTHS})
        st.session_state[MONTHLY_LOADING_KEY] = df

    _nc1, _nc2, _nc3 = st.columns([3, 1, 3], gap="small", vertical_alignment="bottom")
    with _nc1:
        num = st.number_input("Number of Months", min_value=1, max_value=60, value=len(df), step=1)
    with _nc2:
        if st.button("Update", use_container_width=True) and num != len(df):
            months = [f"Month {i+1:02d}" for i in range(num)]
            if num > len(df):
                extra = pd.DataFrame({"Month": months[len(df):],
                                      "Loading Factor (%)": [100.0]*(num-len(df)),
                                      "Weightage Distribution": [100.0/num]*(num-len(df))})
                df = pd.concat([df.iloc[:len(df)], extra], ignore_index=True)
            else:
                df = df.iloc[:num].reset_index(drop=True)
            df["Weightage Distribution"] = 100.0 / num
            st.session_state[MONTHLY_LOADING_KEY] = df; st.rerun()
    with _nc3:
        st.info(f"⚡ Total weightage: {df['Weightage Distribution'].sum():.1f}%")

    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)

    # Table-style month editor
    header = st.columns([2,2,2,2,3], gap="small")
    for h, lbl in zip(header, ["Month","Loading Factor (%)","Weightage (%)","Effective %","Est. Cost"]):
        h.markdown(f"<div style='font-size:0.75rem;font-weight:600;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.05em'>{lbl}</div>", unsafe_allow_html=True)

    for idx, row in df.iterrows():
        cols = st.columns([2,2,2,2,3], gap="small")
        with cols[0]:
            st.markdown(f"<div style='padding:0.5rem 0;font-weight:500'>{row['Month']}</div>", unsafe_allow_html=True)
        with cols[1]:
            df.loc[idx, "Loading Factor (%)"] = st.number_input(
                "lf", min_value=0.0, max_value=200.0, value=float(row["Loading Factor (%)"]),
                step=5.0, format="%.1f", key=f"load_{idx}", label_visibility="collapsed")
        with cols[2]:
            df.loc[idx, "Weightage Distribution"] = st.number_input(
                "wd", min_value=0.0, max_value=100.0, value=float(row["Weightage Distribution"]),
                step=0.1, format="%.1f", key=f"w_{idx}", label_visibility="collapsed")
        with cols[3]:
            eff = (df.loc[idx,"Loading Factor (%)"] / 100) * (df.loc[idx,"Weightage Distribution"] / 100) * 100
            st.markdown(f"<div style='padding:0.5rem 0;font-weight:500;color:var(--teal)'>{eff:.1f}%</div>", unsafe_allow_html=True)
        with cols[4]:
            cost = total_labour * (df.loc[idx,"Weightage Distribution"]/100) * (df.loc[idx,"Loading Factor (%)"]/100)
            st.markdown(f"<div style='padding:0.5rem 0'>{currency} {cost:,.0f}</div>", unsafe_allow_html=True)

    if abs(df["Weightage Distribution"].sum() - 100) > 0.01:
        st.warning(f"⚠️ Weightage totals {df['Weightage Distribution'].sum():.1f}% — should be 100%")

    st.session_state[MONTHLY_LOADING_KEY] = df

    # Preview chart
    breakdown = apply_monthly_loading(labour_df, third_party_df, df, currency)
    if breakdown["months"]:
        preview = pd.DataFrame([{
            "Month": m,
            "Labour": breakdown["monthly_labour"].get(m, 0),
            "Third Party": breakdown["monthly_third_party"].get(m, 0),
        } for m in breakdown["months"]])
        fig = px.bar(preview, x="Month", y=["Labour","Third Party"],
                     color_discrete_map={"Labour": PETRONAS["teal"], "Third Party": PETRONAS["gold"]},
                     template="plotly_dark" if st.session_state["dark_mode"] else "plotly_white",
                     barmode="stack")
        fig.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=0,r=0,t=10,b=0), font=dict(family="DM Sans"),
            legend=dict(orientation="h", y=-0.15)
        )
        st.plotly_chart(fig, use_container_width=True)

    b1, b2, b3 = st.columns(3, gap="small")
    with b1:
        if st.button("← Back", use_container_width=True):
            st.session_state["page"] = "NON_LABOUR"; st.rerun()
    with b2:
        if st.button("Reset", use_container_width=True):
            months = [f"Month {i+1:02d}" for i in range(12)]
            st.session_state[MONTHLY_LOADING_KEY] = pd.DataFrame({
                "Month": months, "Loading Factor (%)": [100.0]*12,
                "Weightage Distribution": [100.0/12]*12
            }); st.rerun()
    with b3:
        if st.button("Next →", use_container_width=True, type="primary"):
            st.session_state["page"] = "TOTALS"; st.rerun()

# ─────────────────────────────────────────────────
# Page: COMPARE (no emojis)
def render_compare():
    st.markdown(f"""
    <div class="petronas-banner fadeup">
        <h1>Project Comparison</h1>
        <p class="subtitle">Compare saved project scenarios side-by-side</p>
    </div>
    """, unsafe_allow_html=True)

    saved = st.session_state.get(SAVED_PROJECTS_KEY, [])
    if not saved:
        st.info("No saved projects yet. Save a project from the Totals or Dashboard page.")
        if st.button("Back to Dashboard"):
            st.session_state["page"] = "SUMMARY"; st.rerun()
        return

    display = [f"{p['name']}  ({p['type_of_schedule']}, {p['type_of_package']}, {p['currency']})" for p in saved]
    col1, col2 = st.columns(2, gap="large")
    with col1:
        i1 = st.selectbox("First Project",  range(len(saved)), format_func=lambda i: display[i], key="c1")
    with col2:
        i2 = st.selectbox("Second Project", range(len(saved)), format_func=lambda i: display[i],
                         index=min(1, len(saved)-1), key="c2")

    if i1 == i2:
        st.warning("Select two different projects to compare.")
    else:
        p1, p2 = saved[i1], saved[i2]
        comp   = compare_projects(p1, p2)

        st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
        section_header("Delta Analysis", "📐")

        m1, m2 = st.columns(2, gap="medium")
        with m1:
            badge = "up" if comp["hours_pct"] > 5 else "down" if comp["hours_pct"] < -5 else "neutral"
            st.markdown(f"""
            <div class="metric-card">
                <div class="label">Manhour Change</div>
                <div class="value">{comp['hours_diff']:+,.0f} hrs</div>
                <div class="sub">
                    <span class="comparison-badge {badge}">{comp['hours_pct']:+.1f}%</span>
                    vs {p1['name']}
                </div>
            </div>""", unsafe_allow_html=True)
        with m2:
            badge2 = "up" if comp["exact_cost_pct"] > 5 else "down" if comp["exact_cost_pct"] < -5 else "neutral"
            st.markdown(f"""
            <div class="metric-card" style="border-top-color:var(--gold)">
                <div class="label">Total Cost Change</div>
                <div class="value">{p2['currency']} {comp['exact_cost_diff']:+,.0f}</div>
                <div class="sub">
                    <span class="comparison-badge {badge2}">{comp['exact_cost_pct']:+.1f}%</span>
                    vs {p1['name']}
                </div>
            </div>""", unsafe_allow_html=True)

        # Side-by-side table
        st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
        compare_df = pd.DataFrame([
            {"Metric":"Schedule",      p1["name"]: p1["type_of_schedule"],   p2["name"]: p2["type_of_schedule"]},
            {"Metric":"Package",       p1["name"]: p1["type_of_package"],    p2["name"]: p2["type_of_package"]},
            {"Metric":"Currency",      p1["name"]: p1["currency"],           p2["name"]: p2["currency"]},
            {"Metric":"Total Hours",   p1["name"]: f"{p1['total_hours']:,.0f}", p2["name"]: f"{p2['total_hours']:,.0f}"},
            {"Metric":"Labour Cost",   p1["name"]: f"{p1['total_labour_exact']:,.2f}", p2["name"]: f"{p2['total_labour_exact']:,.2f}"},
            {"Metric":"Third Party",   p1["name"]: f"{p1['total_third_party']:,.2f}",  p2["name"]: f"{p2['total_third_party']:,.2f}"},
            {"Metric":"Total Bid",     p1["name"]: f"{p1['total_exact']:,.2f}",        p2["name"]: f"{p2['total_exact']:,.2f}"},
        ])
        st.dataframe(compare_df, use_container_width=True, hide_index=True)

    st.markdown("<div class='pdivider'></div>", unsafe_allow_html=True)
    section_header("Saved Projects", "🗄️")
    for i, p in enumerate(saved):
        cols = st.columns([4, 2, 2, 1], gap="small")
        with cols[0]:
            st.markdown(f"<div style='font-weight:600'>{p['name']}</div>"
                        f"<div style='font-size:0.78rem;color:var(--text-muted)'>"
                        f"{p['type_of_schedule']} · {p['type_of_package']} · {p['currency']}</div>",
                        unsafe_allow_html=True)
        with cols[1]:
            st.markdown(f"<div style='font-size:0.85rem;padding-top:0.2rem'>{p['currency']} {p['total_exact']:,.0f}</div>", unsafe_allow_html=True)
        with cols[2]:
            ts = datetime.fromisoformat(p["timestamp"]).strftime("%Y-%m-%d %H:%M")
            st.markdown(f"<div style='font-size:0.78rem;color:var(--text-muted);padding-top:0.3rem'>{ts}</div>", unsafe_allow_html=True)
        with cols[3]:
            if st.button("🗑️", key=f"del_{i}"):
                saved.pop(i)
                st.session_state[SAVED_PROJECTS_KEY] = saved; st.rerun()

    col1, col2 = st.columns(2, gap="small")
    with col1:
        if st.button("Clear All", use_container_width=True):
            st.session_state[SAVED_PROJECTS_KEY] = []; st.rerun()
    with col2:
        if st.button("Back to Dashboard", use_container_width=True, type="primary"):
            st.session_state["page"] = "SUMMARY"; st.rerun()

# ─────────────────────────────────────────────────
# App header (no emojis)
st.markdown(f"""
<div style="display:flex;align-items:center;gap:1rem;margin-bottom:0.5rem">
    <div>{_logo_html}</div>
    <div>
        <span style="font-weight:700;font-size:1.1rem;color:var(--teal);letter-spacing:-0.02em">MEC TOOL</span>
        <span style="color:{'rgba(255,255,255,0.6)' if st.session_state.get('dark_mode', False) else 'rgba(10,22,40,0.55)'};font-size:0.85rem;margin-left:0.75rem">Major Engineering Contract · PETRONAS Upstream CE</span>
    </div>
</div>
""", unsafe_allow_html=True)

render_navigation()

# ─────────────────────────────────────────────────
# Guard: show upload prompt if MEC.csv not yet loaded
if not st.session_state["mec_data_loaded"]:
    st.markdown("""
    <div class="petronas-banner fadeup">
        <h1>Welcome to MEC Tool</h1>
        <p class="subtitle">Major Engineering Contract Cost Estimation — PETRONAS Upstream CE</p>
    </div>
    """, unsafe_allow_html=True)
    st.info("← Please upload **MEC.csv** in the sidebar to get started. Use the **›** arrow at the top-left to open the sidebar if it is collapsed.")
else:
    # Router
    page = st.session_state["page"]
    if   page == "MAIN":        render_main()
    elif page == "TABLE":       render_table()
    elif page == "THIRD_PARTY": render_third_party()
    elif page == "NON_LABOUR":  render_non_labour()
    elif page == "LOADING":     render_loading()
    elif page == "TOTALS":      render_totals()
    elif page == "SUMMARY":     render_summary()
    elif page == "COMPARE":     render_compare()
    else:
        st.session_state["page"] = "MAIN"
        st.rerun()


